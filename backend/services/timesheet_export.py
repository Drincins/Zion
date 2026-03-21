from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Optional, Sequence
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy import and_
from sqlalchemy.orm import Session, joinedload

from backend.bd.models import Attendance, Position, User
from backend.services.payroll_export import _autosize_sheet, _style_sheet, FIRED_FILL, TOTAL_FILL


def _date_range(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def _format_time(value) -> str:
    if not value:
        return ""
    return value.strftime("%H:%M")


def _format_shift(attendance: Attendance) -> str:
    start = _format_time(attendance.open_time)
    end = _format_time(attendance.close_time)
    if start and end:
        return f"{start}-{end}"
    if start:
        return f"{start}-"
    return ""


def _minutes_to_hours(minutes: int) -> float:
    return round((minutes or 0) / 60.0, 2)

def _weekday_label(day: date) -> str:
    labels = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    return labels[day.weekday()]


def _safe_sheet_title(title: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "", title or "").strip()
    if not cleaned:
        cleaned = "Сотрудник"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 1
    while candidate in existing:
        suffix = f" {counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing.add(candidate)
    return candidate


def build_timesheet_report(
    db: Session,
    *,
    date_from: date,
    date_to: date,
    restaurant_id: int,
    restaurant_subdivision_ids: Optional[Sequence[int]] = None,
    position_ids: Optional[Sequence[int]] = None,
) -> BytesIO:
    days = _date_range(date_from, date_to)

    attendance_map: dict[int, dict[date, list[str]]] = {}
    totals: dict[int, dict[str, int]] = {}
    daily_minutes: dict[int, dict[date, int]] = {}
    daily_night: dict[int, dict[date, int]] = {}

    attendance_query = (
        db.query(Attendance)
        .filter(
            Attendance.restaurant_id == restaurant_id,
            Attendance.close_date.isnot(None),
            and_(Attendance.open_date >= date_from, Attendance.open_date <= date_to),
        )
        .order_by(Attendance.user_id.asc(), Attendance.open_date.asc(), Attendance.open_time.asc())
    )
    if position_ids:
        attendance_query = attendance_query.filter(
            Attendance.position_id.in_([int(pid) for pid in position_ids])
        )
    if restaurant_subdivision_ids:
        attendance_query = attendance_query.join(Position, Attendance.position_id == Position.id)
        attendance_query = attendance_query.filter(
            Position.restaurant_subdivision_id.in_([int(sid) for sid in restaurant_subdivision_ids])
        )

    attendance_rows = attendance_query.all()

    for attendance in attendance_rows:
        shifts = attendance_map.setdefault(attendance.user_id, {})
        day_shifts = shifts.setdefault(attendance.open_date, [])
        shift_value = _format_shift(attendance)
        if shift_value:
            day_shifts.append(shift_value)
        summary = totals.setdefault(attendance.user_id, {"minutes": 0, "night": 0})
        minutes_value = int(attendance.duration_minutes or 0)
        night_value = int(attendance.night_minutes or 0)
        summary["minutes"] += minutes_value
        summary["night"] += night_value
        minutes_by_day = daily_minutes.setdefault(attendance.user_id, {})
        night_by_day = daily_night.setdefault(attendance.user_id, {})
        minutes_by_day[attendance.open_date] = minutes_by_day.get(attendance.open_date, 0) + minutes_value
        night_by_day[attendance.open_date] = night_by_day.get(attendance.open_date, 0) + night_value

    # Only include employees who have shifts in the selected period.
    user_ids = set(attendance_map.keys())

    users = []
    if user_ids:
        users = (
            db.query(User)
            .options(joinedload(User.position).joinedload(Position.restaurant_subdivision))
            .filter(User.id.in_(list(user_ids)))
            .all()
        )
        users.sort(
            key=lambda user: (
                (user.last_name or "").lower(),
                (user.first_name or "").lower(),
                (user.username or "").lower(),
            )
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    headers = ["Фамилия Имя", "Должность"]
    headers.extend(day.strftime("%d.%m") for day in days)
    headers.extend(["Итого часов", "Итого ночных часов"])
    ws.append(headers)

    total_hours_col = 2 + len(days) + 1
    total_night_col = total_hours_col + 1

    for user in users:
        full_name = f"{user.last_name or ''} {user.first_name or ''}".strip() or user.username or ""
        position_name = user.position.name if user.position else ""
        row = [full_name, position_name]

        shifts_by_day = attendance_map.get(user.id, {})
        for day in days:
            row.append("\n".join(shifts_by_day.get(day, [])))

        summary = totals.get(user.id, {"minutes": 0, "night": 0})
        row.append(_minutes_to_hours(summary["minutes"]))
        row.append(_minutes_to_hours(summary["night"]))
        ws.append(row)

        row_idx = ws.max_row
        for col_idx in range(3, 3 + len(days)):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_idx, column=total_hours_col).number_format = "0.00"
        ws.cell(row=row_idx, column=total_night_col).number_format = "0.00"
        if getattr(user, "fired", False):
            for cell in ws[row_idx]:
                cell.fill = FIRED_FILL

    if users:
        total_minutes = sum((summary.get("minutes", 0) for summary in totals.values()), 0)
        total_night = sum((summary.get("night", 0) for summary in totals.values()), 0)
        total_row = ["Итого", ""] + [""] * len(days)
        total_row.append(_minutes_to_hours(total_minutes))
        total_row.append(_minutes_to_hours(total_night))
        ws.append(total_row)
        total_row_idx = ws.max_row
        ws.cell(row=total_row_idx, column=total_hours_col).number_format = "0.00"
        ws.cell(row=total_row_idx, column=total_night_col).number_format = "0.00"
        for cell in ws[total_row_idx]:
            cell.fill = TOTAL_FILL

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).coordinate}"

    _style_sheet(ws, freeze="C2")
    _autosize_sheet(ws, min_width=10.0, max_width=30.0)

    existing_titles = {ws.title}
    for user in users:
        full_name = f"{user.last_name or ''} {user.first_name or ''}".strip() or user.username or ""
        sheet_title = _safe_sheet_title(full_name, existing_titles)
        user_ws = wb.create_sheet(title=sheet_title)

        user_ws.append(["Дата", "День", "Смены", "Часы", "Ночные"])
        shifts_by_day = attendance_map.get(user.id, {})
        minutes_by_day = daily_minutes.get(user.id, {})
        night_by_day = daily_night.get(user.id, {})

        for day in days:
            shifts = "\n".join(shifts_by_day.get(day, []))
            minutes_value = minutes_by_day.get(day, 0)
            night_value = night_by_day.get(day, 0)
            user_ws.append(
                [
                    day.strftime("%d.%m.%Y"),
                    _weekday_label(day),
                    shifts,
                    _minutes_to_hours(minutes_value) if minutes_value else "",
                    _minutes_to_hours(night_value) if night_value else "",
                ]
            )

        summary = totals.get(user.id, {"minutes": 0, "night": 0})
        user_ws.append(
            [
                "Итого",
                "",
                "",
                _minutes_to_hours(summary["minutes"]),
                _minutes_to_hours(summary["night"]),
            ]
        )
        total_row_idx = user_ws.max_row
        for cell in user_ws[total_row_idx]:
            cell.fill = TOTAL_FILL

        for row in user_ws.iter_rows(min_row=2, max_row=user_ws.max_row, min_col=1, max_col=5):
            row[2].alignment = Alignment(wrap_text=True, vertical="top")

        _style_sheet(user_ws, freeze="A2")
        _autosize_sheet(user_ws, min_width=10.0, max_width=32.0)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
