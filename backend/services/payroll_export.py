from __future__ import annotations

import calendar
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from types import SimpleNamespace
from typing import Iterable, Mapping, Optional, Sequence

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

from backend.bd.iiko_catalog import IikoPaymentMethod
from backend.bd.iiko_sales import IikoSaleItem, IikoSaleOrder
from backend.bd.models import (
    Attendance,
    Company,
    LaborSummarySettings,
    PayrollAdvanceStatement,
    PayrollAdjustment,
    PayrollAdjustmentType,
    Position,
    Restaurant,
    User,
    PaymentFormat,
)
from backend.services.payroll_calculations import calc_amounts as calc_payroll_amounts
from backend.services.payroll_calculations import resolve_rate as resolve_payroll_rate


MONEY_QUANT = Decimal("0.01")
NUMBER_FORMAT_CURRENCY = '#,##0.00 "₽"'
NUMBER_FORMAT_NUMBER = "0.00"
NUMBER_FORMAT_PERCENT = "0.00%"

HEADER_FILL = PatternFill(fill_type="solid", start_color="FFF7E5D6", end_color="FFF7E5D6")
TOTAL_FILL = PatternFill(fill_type="solid", start_color="FFFDF3C4", end_color="FFFDF3C4")
NEG_TOTAL_FILL = PatternFill(fill_type="solid", start_color="FFF9D6D5", end_color="FFF9D6D5")
SPACER_FILL = PatternFill(fill_type="solid", start_color="FFF0F0F0", end_color="FFF0F0F0")
FIRED_FILL = PatternFill(fill_type="solid", start_color="FFF7E5E5", end_color="FFF7E5E5")
SUBDIVISION_TOTAL_FILL = PatternFill(fill_type="solid", start_color="FFE8F1FF", end_color="FFE8F1FF")
SUBDIVISION_TOTAL_FONT = Font(bold=True, size=12)
KEY_DEDUCTION_NAMES = (
    "Аванс (Карта)",
    "Зарплата (Карта)",
)
_KEY_DEDUCTION_NORMALIZE_RE = re.compile(r"[^0-9a-zа-яё]+")
THIN_BORDER = Border(
    left=Side(style="thin", color="FFB8B8B8"),
    right=Side(style="thin", color="FFB8B8B8"),
    top=Side(style="thin", color="FFB8B8B8"),
    bottom=Side(style="thin", color="FFB8B8B8"),
)
INVALID_SHEET_CHARS = set("[]:*?/\\")
REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT = "sum_without_discount"
REVENUE_AMOUNT_MODE_SUM_WITH_DISCOUNT = "sum_with_discount"
REVENUE_AMOUNT_MODE_DISCOUNT_ONLY = "discount_only"


def _resolve_key_deduction_name(name: Optional[str]) -> Optional[str]:
    """Return canonical key deduction label for a given adjustment name."""
    if not name:
        return None

    # Real-world data uses different labels for the same "card payout" deductions
    # (e.g. "Авансы карты", "З/П (карта)"), so we normalize and detect by tokens.
    compact = _KEY_DEDUCTION_NORMALIZE_RE.sub("", name.strip().lower())
    if not compact:
        return None

    if "карт" not in compact:
        return None

    if any(token in compact for token in ("аванс", "аванас", "аваны")):
        return KEY_DEDUCTION_NAMES[0]

    if any(token in compact for token in ("зп", "зарп", "заработ")):
        return KEY_DEDUCTION_NAMES[1]
    return None


def _sanitize_sheet_title(title: Optional[str]) -> str:
    raw = (title or "").strip() or "Ведомость"
    cleaned = "".join(" " if ch in INVALID_SHEET_CHARS else ch for ch in raw)
    cleaned = " ".join(cleaned.split())
    return cleaned[:31] or "Ведомость"


def _unique_sheet_title(base: Optional[str], used: set[str]) -> str:
    title = _sanitize_sheet_title(base)
    if title not in used:
        used.add(title)
        return title
    idx = 2
    while True:
        suffix = f" ({idx})"
        max_len = max(1, 31 - len(suffix))
        candidate = (title[:max_len].rstrip() or "Ведомость") + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
        idx += 1


def _quantize(amount: Decimal) -> Decimal:
    return amount.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _decimal(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _normalize_statement_kind(value: Optional[str]) -> Optional[str]:
    kind = str(value or "").strip().lower()
    if kind == "salary":
        return "salary"
    if kind == "advance":
        return "advance"
    return None


def _clean_optional_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_revenue_amount_mode(value: Optional[str]) -> str:
    clean = str(value or "").strip().lower()
    aliases = {
        "": REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT,
        "sum": REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT,
        "gross": REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT,
        REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT: REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT,
        REVENUE_AMOUNT_MODE_SUM_WITH_DISCOUNT: REVENUE_AMOUNT_MODE_SUM_WITH_DISCOUNT,
        "net": REVENUE_AMOUNT_MODE_SUM_WITH_DISCOUNT,
        REVENUE_AMOUNT_MODE_DISCOUNT_ONLY: REVENUE_AMOUNT_MODE_DISCOUNT_ONLY,
        "discount": REVENUE_AMOUNT_MODE_DISCOUNT_ONLY,
    }
    return aliases.get(clean, REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT)


def _is_not_deleted_expr(field_expr):
    normalized = func.upper(func.trim(func.coalesce(field_expr, "")))
    return or_(normalized == "", normalized == "NOT_DELETED")


def _resolve_revenue_sum_expr(amount_mode: str):
    discount_abs_expr = func.abs(func.coalesce(IikoSaleItem.discount_sum, 0))
    if amount_mode == REVENUE_AMOUNT_MODE_DISCOUNT_ONLY:
        return discount_abs_expr
    if amount_mode == REVENUE_AMOUNT_MODE_SUM_WITH_DISCOUNT:
        return func.coalesce(IikoSaleItem.sum, 0) - discount_abs_expr
    return func.coalesce(IikoSaleItem.sum, 0)


def _load_real_money_payment_method_maps(
    db: Session,
    *,
    company_id: Optional[int],
) -> tuple[set[str], dict[str, str], int]:
    query = db.query(
        IikoPaymentMethod.guid,
        IikoPaymentMethod.name,
        IikoPaymentMethod.category,
        IikoPaymentMethod.is_active,
    )
    if company_id is not None:
        query = query.filter(
            or_(
                IikoPaymentMethod.company_id == company_id,
                IikoPaymentMethod.company_id.is_(None),
            )
        )
    rows = query.all()

    real_money_guids: set[str] = set()
    guid_by_name: dict[str, str] = {}
    categorized_count = 0
    for guid, name, category, is_active in rows:
        clean_guid = _clean_optional_text(guid)
        clean_name = _clean_optional_text(name)
        clean_category = (_clean_optional_text(category) or "").casefold()
        if not clean_guid:
            continue
        clean_guid_norm = clean_guid.casefold()
        if clean_name:
            guid_by_name.setdefault(clean_name.casefold(), clean_guid_norm)
        if clean_category:
            categorized_count += 1
        if is_active is False:
            continue
        if clean_category == "real_money":
            real_money_guids.add(clean_guid_norm)
    return real_money_guids, guid_by_name, categorized_count


def _load_revenue_settings_for_company(
    db: Session,
    *,
    company_id: Optional[int],
) -> tuple[bool, bool, str]:
    if company_id is None:
        return True, True, REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT

    entry = (
        db.query(LaborSummarySettings)
        .filter(LaborSummarySettings.company_id == company_id)
        .first()
    )
    if not entry:
        return True, True, REVENUE_AMOUNT_MODE_SUM_WITHOUT_DISCOUNT

    return (
        bool(entry.revenue_exclude_deleted),
        bool(entry.revenue_real_money_only),
        _normalize_revenue_amount_mode(entry.revenue_amount_mode),
    )


def _calc_revenue_amount(
    db: Session,
    *,
    company_id: Optional[int],
    restaurant_ids: Sequence[int],
    date_from: date,
    date_to: date,
    exclude_deleted_orders: bool,
    real_money_only: bool,
    amount_mode: str,
) -> Decimal:
    normalized_restaurant_ids: set[int] = set()
    for raw in restaurant_ids or []:
        try:
            parsed = int(raw)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            normalized_restaurant_ids.add(parsed)
    normalized_restaurant_ids = sorted(normalized_restaurant_ids)
    if not normalized_restaurant_ids:
        return Decimal("0")

    amount_expr = _resolve_revenue_sum_expr(amount_mode)
    base_q = (
        db.query(IikoSaleItem)
        .join(IikoSaleOrder, IikoSaleOrder.id == IikoSaleItem.order_id)
        .filter(IikoSaleOrder.restaurant_id.in_(normalized_restaurant_ids))
        .filter(IikoSaleOrder.open_date >= date_from, IikoSaleOrder.open_date <= date_to)
    )
    if exclude_deleted_orders:
        is_not_deleted = _is_not_deleted_expr(
            func.coalesce(
                IikoSaleItem.raw_payload["OrderDeleted"].astext,
                IikoSaleOrder.raw_payload["OrderDeleted"].astext,
            )
        ) & _is_not_deleted_expr(
            func.coalesce(
                IikoSaleItem.raw_payload["DeletedWithWriteoff"].astext,
                IikoSaleOrder.raw_payload["DeletedWithWriteoff"].astext,
            )
        )
        base_q = base_q.filter(is_not_deleted)

    if not real_money_only:
        value = base_q.with_entities(func.coalesce(func.sum(amount_expr), 0)).scalar()
        return _decimal(value)

    non_cash_id_expr = func.trim(func.coalesce(IikoSaleItem.raw_payload["NonCashPaymentType.Id"].astext, ""))
    non_cash_name_expr = func.trim(func.coalesce(IikoSaleItem.raw_payload["NonCashPaymentType"].astext, ""))
    base_q = base_q.filter(non_cash_id_expr == "").filter(non_cash_name_expr == "")

    payment_guid_expr = func.coalesce(
        IikoSaleItem.raw_payload["PayTypes.GUID"].astext,
        IikoSaleItem.raw_payload["PaymentType.Id"].astext,
        IikoSaleItem.raw_payload["NonCashPaymentType.Id"].astext,
    )
    payment_guid_expr_norm = func.lower(func.trim(func.coalesce(payment_guid_expr, "")))
    payment_name_expr = func.coalesce(
        IikoSaleItem.raw_payload["PayTypes"].astext,
        IikoSaleItem.raw_payload["PaymentType"].astext,
        IikoSaleItem.raw_payload["NonCashPaymentType"].astext,
    )

    grouped_rows = (
        base_q.with_entities(
            payment_guid_expr_norm.label("payment_guid_norm"),
            payment_name_expr.label("payment_name"),
            func.coalesce(func.sum(amount_expr), 0).label("amount"),
        )
        .group_by(payment_guid_expr_norm, payment_name_expr)
        .all()
    )
    if not grouped_rows:
        return Decimal("0")

    real_money_guids, guid_by_name, categorized_count = _load_real_money_payment_method_maps(
        db,
        company_id=company_id,
    )

    if real_money_guids:
        total = Decimal("0")
        for row in grouped_rows:
            raw_guid_norm = _clean_optional_text(getattr(row, "payment_guid_norm", None))
            raw_name = _clean_optional_text(getattr(row, "payment_name", None))
            resolved_guid_norm = raw_guid_norm.casefold() if raw_guid_norm else None
            if not resolved_guid_norm and raw_name:
                resolved_guid_norm = guid_by_name.get(raw_name.casefold())
            if not resolved_guid_norm or resolved_guid_norm not in real_money_guids:
                continue
            total += _decimal(getattr(row, "amount", 0))
        return total

    if categorized_count > 0:
        return Decimal("0")

    total = Decimal("0")
    for row in grouped_rows:
        total += _decimal(getattr(row, "amount", 0))
    return total


def _resolve_revenue_amount_for_report(
    db: Session,
    *,
    company_id: Optional[int],
    restaurant_ids: Sequence[int],
    date_from: Optional[date],
    date_to: Optional[date],
) -> Optional[Decimal]:
    if not date_from or not date_to:
        return None

    normalized_restaurant_ids: set[int] = set()
    for raw in restaurant_ids or []:
        try:
            parsed = int(raw)
        except (TypeError, ValueError):
            continue
        if parsed > 0:
            normalized_restaurant_ids.add(parsed)
    normalized_restaurant_ids = sorted(normalized_restaurant_ids)
    if not normalized_restaurant_ids:
        return None

    revenue_exclude_deleted, revenue_real_money_only, revenue_amount_mode = _load_revenue_settings_for_company(
        db,
        company_id=company_id,
    )
    revenue_amount = _calc_revenue_amount(
        db,
        company_id=company_id,
        restaurant_ids=normalized_restaurant_ids,
        date_from=date_from,
        date_to=date_to,
        exclude_deleted_orders=revenue_exclude_deleted,
        real_money_only=revenue_real_money_only,
        amount_mode=revenue_amount_mode,
    )

    if len(normalized_restaurant_ids) == 1 and revenue_amount <= 0 and company_id is not None:
        company_restaurant_ids = [
            int(item[0])
            for item in (
                db.query(Restaurant.id).filter(Restaurant.company_id == company_id).all()
            )
            if item and item[0] is not None
        ]
        if company_restaurant_ids:
            company_revenue_amount = _calc_revenue_amount(
                db,
                company_id=company_id,
                restaurant_ids=company_restaurant_ids,
                date_from=date_from,
                date_to=date_to,
                exclude_deleted_orders=revenue_exclude_deleted,
                real_money_only=revenue_real_money_only,
                amount_mode=revenue_amount_mode,
            )
            if company_revenue_amount > 0:
                revenue_amount = company_revenue_amount

    return _quantize(revenue_amount)


def _resolve_rate(
    user: User,
    position: Optional[Position],
    payment_mode: Optional[str],
    *,
    stats_rate: Optional[Decimal] = None,
) -> Decimal:
    return resolve_payroll_rate(user, position, payment_mode, stats_rate=stats_rate)


def _period_bounds(period: str) -> tuple[date, date]:
    """Parse YYYY-MM into first/last day of month."""
    year, month = map(int, period.split("-"))
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    return first_day, last_day


def resolve_payroll_export_bounds(
    *,
    period: Optional[str],
    date_from: Optional[date],
    date_to: Optional[date],
) -> tuple[date, date]:
    """Resolve export date range from explicit dates or period."""
    if date_from or date_to:
        start = date_from or date_to
        end = date_to or date_from
        if not start or not end:
            raise ValueError("date_from/date_to must be provided")
        if end < start:
            raise ValueError("date_to must be >= date_from")
        return start, end
    if period:
        return _period_bounds(period)
    raise ValueError("period or date_from/date_to is required")


@dataclass
class UserPayrollRow:
    user: User
    company: Optional[Company]
    restaurant: Optional[Restaurant]
    subdivision_name: Optional[str]
    position: Optional[Position]
    rate: Decimal
    fact_hours: Decimal
    fact_shifts: Optional[Decimal]
    night_hours: Decimal
    base_amount: Decimal
    night_bonus: Decimal
    hours_total: Decimal
    adjustments_by_type: dict[int, Decimal]
    total_adjustments: Decimal
    total_accruals: Decimal
    total_deductions: Decimal
    total_to_pay: Decimal
    adjustments_only: bool = False


@dataclass
class AdjustmentSnapshotRow:
    user_id: int
    adjustment_type_id: int
    amount: Decimal
    date: Optional[date]
    comment: Optional[str]
    restaurant_id: Optional[int] = None
    restaurant_name: Optional[str] = None
    subdivision_name: Optional[str] = None
    statement_id: Optional[int] = None


@dataclass
class PayrollReportSnapshot:
    results: list[UserPayrollRow]
    adj_type_map: dict[int, PayrollAdjustmentType]
    adj_rows: list[AdjustmentSnapshotRow]
    user_map: dict[int, User]
    salary_factor: Decimal
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    company_id: Optional[int] = None
    restaurant_id: Optional[int] = None
    statement_kind: Optional[str] = None


def _calc_amounts(
    *,
    rate: Decimal,
    payment_mode: Optional[str],
    fact_hours: Decimal,
    hours_per_shift: Optional[Decimal],
    monthly_shift_norm: Optional[Decimal],
    night_hours: Decimal,
    night_bonus_enabled: bool,
    night_bonus_percent: Decimal,
    salary_factor: Decimal,
) -> tuple[Decimal, Optional[Decimal], Decimal, Decimal]:
    return calc_payroll_amounts(
        rate=rate,
        payment_mode=payment_mode,
        fact_hours=fact_hours,
        hours_per_shift=hours_per_shift,
        monthly_shift_norm=monthly_shift_norm,
        night_hours=night_hours,
        night_bonus_enabled=night_bonus_enabled,
        night_bonus_percent=night_bonus_percent,
        salary_factor=salary_factor,
    )


def _collect_attendance(
    db: Session,
    start: date,
    end: date,
    *,
    company_id: Optional[int],
    restaurant_id: Optional[int],
    subdivision_id: Optional[int],
    user_ids: Optional[Sequence[int]] = None,
) -> dict[int, list[dict]]:
    q = (
        db.query(
            Attendance.user_id,
            Attendance.position_id,
            Attendance.rate.label("rate_value"),
            func.sum(Attendance.duration_minutes).label("duration_sum"),
            func.sum(Attendance.night_minutes).label("night_sum"),
            func.max(Attendance.restaurant_id).label("restaurant_id"),
        )
        .filter(Attendance.close_date.isnot(None))
        # ??? ???? ??????? ? ???? ???????? ?????, ???? ???? ????????? ? ?????? ????
        .filter(Attendance.open_date >= start, Attendance.open_date <= end)
    )

    if restaurant_id:
        q = q.filter(Attendance.restaurant_id == restaurant_id)
    if company_id:
        q = q.join(Restaurant, Attendance.restaurant_id == Restaurant.id)
        q = q.filter(Restaurant.company_id == company_id)
    if subdivision_id:
        q = q.join(Position, Attendance.position_id == Position.id)
        q = q.filter(Position.restaurant_subdivision_id == subdivision_id)

    if user_ids is not None:
        ids = [int(uid) for uid in user_ids]
        if not ids:
            return {}
        q = q.filter(Attendance.user_id.in_(ids))
    rows = q.group_by(Attendance.user_id, Attendance.position_id, Attendance.rate).all()
    data: dict[int, list[dict]] = {}
    for row in rows:
        data.setdefault(row.user_id, []).append(
            {
                "position_id": row.position_id,
                "rate": _decimal(row.rate_value) if row.rate_value is not None else None,
                "duration_minutes": int(row.duration_sum or 0),
                "night_minutes": int(row.night_sum or 0),
                "restaurant_id": row.restaurant_id,
            }
        )
    return data


def _collect_adjustments(
    db: Session,
    start: date,
    end: date,
    user_ids: Optional[Iterable[int]],
    *,
    company_id: Optional[int],
    restaurant_id: Optional[int],
    subdivision_id: Optional[int],
) -> tuple[dict[int, dict[int, Decimal]], dict[int, PayrollAdjustmentType], list[PayrollAdjustment]]:
    ids = None
    if user_ids is not None:
        ids = [int(uid) for uid in user_ids]
        if not ids:
            return {}, {}, []

    q = (
        db.query(PayrollAdjustment, PayrollAdjustmentType)
        .join(PayrollAdjustmentType, PayrollAdjustment.adjustment_type_id == PayrollAdjustmentType.id)
        .join(User, PayrollAdjustment.user_id == User.id)
        .filter(PayrollAdjustment.date >= start, PayrollAdjustment.date <= end)
    )
    if ids is not None:
        q = q.filter(PayrollAdjustment.user_id.in_(ids))
    if company_id:
        q = q.filter(User.company_id == company_id)
    if restaurant_id:
        q = q.filter(PayrollAdjustment.restaurant_id == restaurant_id)
    if subdivision_id:
        q = q.join(Position, Position.id == User.position_id)
        q = q.filter(Position.restaurant_subdivision_id == subdivision_id)

    by_user: dict[int, dict[int, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    type_map: dict[int, PayrollAdjustmentType] = {}
    adjustments_list: list[PayrollAdjustment] = []
    for adj, adj_type in q.all():
        signed = _decimal(adj.amount)
        # Amounts for deductions are already stored with a minus sign; only flip
        # if legacy data has a positive value to avoid inflating totals.
        if adj_type.kind == "deduction" and signed > 0:
            signed = signed * Decimal("-1")
        by_user[adj.user_id][adj.adjustment_type_id] += signed
        type_map[adj_type.id] = adj_type
        adjustments_list.append(adj)
    return by_user, type_map, adjustments_list


def _style_sheet(ws, *, freeze: str = "A2") -> None:
    """Apply header fill, grid borders, and freeze pane."""
    if ws.max_row == 0 or ws.max_column == 0:
        return

    header_row = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column))
    for cell in header_row:
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = THIN_BORDER

    ws.freeze_panes = freeze


def _apply_number_formats(ws, *, currency_cols: set[int], number_cols: set[int], date_cols: Optional[set[int]] = None) -> None:
    """Assign number formats to target columns across all data rows."""
    if ws.max_row < 2 or ws.max_column == 0:
        return
    date_cols = date_cols or set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.col_idx in currency_cols:
                cell.number_format = NUMBER_FORMAT_CURRENCY
            elif cell.col_idx in number_cols:
                cell.number_format = NUMBER_FORMAT_NUMBER
            elif cell.col_idx in date_cols:
                cell.number_format = "yyyy-mm-dd"


def _autosize_sheet(ws, *, min_width: float = 10.0, max_width: float = 50.0) -> None:
    """Resize columns based on content width for better readability."""
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column if col_cells else None
        if col_idx is None:
            continue
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        width = min(max_width, max(min_width, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_payroll_row_formulas(
    ws,
    row_idx: int,
    *,
    base_amount_col: int,
    night_bonus_col: int,
    total_hours_col: int,
    adj_start_col: int,
    adj_end_col: int,
    tail_start_col: int,
) -> None:
    """Apply formulas for totals on a payroll row."""
    base_letter = get_column_letter(base_amount_col)
    night_letter = get_column_letter(night_bonus_col)
    total_letter = get_column_letter(total_hours_col)
    ws.cell(row=row_idx, column=total_hours_col).value = f"={base_letter}{row_idx}+{night_letter}{row_idx}"

    accruals_col = tail_start_col
    deductions_col = tail_start_col + 1
    total_to_pay_col = tail_start_col + 2

    if adj_end_col >= adj_start_col:
        adj_start_letter = get_column_letter(adj_start_col)
        adj_end_letter = get_column_letter(adj_end_col)
        adj_range = f"{adj_start_letter}{row_idx}:{adj_end_letter}{row_idx}"
        ws.cell(row=row_idx, column=accruals_col).value = f"=SUMIF({adj_range},\">0\")"
        ws.cell(row=row_idx, column=deductions_col).value = f"=-SUMIF({adj_range},\"<0\")"
        ws.cell(row=row_idx, column=total_to_pay_col).value = f"={total_letter}{row_idx}+SUM({adj_range})"
    else:
        ws.cell(row=row_idx, column=accruals_col).value = "=0"
        ws.cell(row=row_idx, column=deductions_col).value = "=0"
        ws.cell(row=row_idx, column=total_to_pay_col).value = f"={total_letter}{row_idx}"


def _apply_payroll_total_formulas(
    ws,
    total_row_idx: int,
    data_end_row: int,
    *,
    base_amount_col: int,
    night_bonus_col: int,
    total_hours_col: int,
    accrual_start_col: int,
    accrual_end_col: int,
    accrual_total_col: int,
    accrued_col: int,
    deduction_start_col: int,
    deduction_end_col: int,
    deduction_total_col: int,
    total_to_pay_col: int,
) -> None:
    """Apply summary formulas for the totals row."""
    if data_end_row < 2:
        return

    def _sum_col(col_idx: int) -> str:
        letter = get_column_letter(col_idx)
        return f"=SUM({letter}2:{letter}{data_end_row})"

    ws.cell(row=total_row_idx, column=base_amount_col).value = _sum_col(base_amount_col)
    ws.cell(row=total_row_idx, column=night_bonus_col).value = _sum_col(night_bonus_col)
    ws.cell(row=total_row_idx, column=total_hours_col).value = _sum_col(total_hours_col)

    if accrual_end_col >= accrual_start_col:
        for col_idx in range(accrual_start_col, accrual_end_col + 1):
            ws.cell(row=total_row_idx, column=col_idx).value = _sum_col(col_idx)
    ws.cell(row=total_row_idx, column=accrual_total_col).value = _sum_col(accrual_total_col)
    ws.cell(row=total_row_idx, column=accrued_col).value = _sum_col(accrued_col)

    if deduction_end_col >= deduction_start_col:
        for col_idx in range(deduction_start_col, deduction_end_col + 1):
            ws.cell(row=total_row_idx, column=col_idx).value = _sum_col(col_idx)
    ws.cell(row=total_row_idx, column=deduction_total_col).value = _sum_col(deduction_total_col)
    ws.cell(row=total_row_idx, column=total_to_pay_col).value = _sum_col(total_to_pay_col)


def _payroll_row_sort_key(row: UserPayrollRow) -> tuple[str, str, str, str]:
    pos_name = getattr(row.position, "name", None) or ""
    last = row.user.last_name or ""
    first = row.user.first_name or ""
    username = row.user.username or ""
    return (pos_name.lower(), last.lower(), first.lower(), username.lower())


def _build_payroll_sheet(
    ws,
    *,
    results: Sequence[UserPayrollRow],
    adj_type_map: Mapping[int, PayrollAdjustmentType],
    salary_factor: Decimal,
    overrides: Optional[dict[int, object]] = None,
    statement_kind: Optional[str] = None,
) -> dict[str, dict[str, Decimal]]:
    rows = list(results or [])
    rows.sort(key=_payroll_row_sort_key)
    statement_kind_code = _normalize_statement_kind(statement_kind)

    def _type_sort_key(type_id: int) -> tuple[int, str]:
        adj_type = adj_type_map[type_id]
        name_lower = (getattr(adj_type, "name", None) or "").lower()
        if getattr(adj_type, "kind", None) == "deduction" and _resolve_key_deduction_name(
            getattr(adj_type, "name", None)
        ):
            return (0, name_lower)
        return (1, name_lower)

    used_type_ids = sorted(adj_type_map.keys(), key=_type_sort_key)
    visible_type_ids = [
        t_id for t_id in used_type_ids if bool(getattr(adj_type_map.get(t_id), "show_in_report", False))
    ]

    headers = [
        "Ресторан",
        "Подразделение",
        "Должность",
        "ФИО",
        "Формат оплаты",
        "Ставка",
        "Часы",
        "Смены",
        "Ночные часы",
        "Сумма за часы",
        "Надбавка за ночь",
        "Итого за часы",
    ]

    used_accrual_type_ids = [t_id for t_id in visible_type_ids if adj_type_map[t_id].kind == "accrual"]
    used_deduction_type_ids = [t_id for t_id in visible_type_ids if adj_type_map[t_id].kind == "deduction"]
    key_deduction_type_ids: dict[int, str] = {}
    for t_id in used_deduction_type_ids:
        adj_type = adj_type_map.get(t_id)
        if not adj_type:
            continue
        canonical = _resolve_key_deduction_name(getattr(adj_type, "name", None))
        if canonical:
            key_deduction_type_ids[t_id] = canonical

    accrual_headers = [f"{adj_type_map[t_id].name} (Начисление)" for t_id in used_accrual_type_ids]
    deduction_headers = [f"{adj_type_map[t_id].name} (Удержание)" for t_id in used_deduction_type_ids]

    all_headers = (
        headers
        + accrual_headers
        + ["Итого начислений", "Итого начислено"]
        + deduction_headers
        + ["Итого удержаний", "К выплате"]
    )
    ws.append(all_headers)

    base_amount_col = headers.index("Сумма за часы") + 1
    night_bonus_col = headers.index("Надбавка за ночь") + 1
    total_hours_col = headers.index("Итого за часы") + 1
    employee_col = headers.index("ФИО") + 1

    accrual_start_col = len(headers) + 1
    accrual_end_col = (
        accrual_start_col + len(used_accrual_type_ids) - 1 if used_accrual_type_ids else accrual_start_col - 1
    )
    accrual_total_col = accrual_end_col + 1
    accrued_col = accrual_total_col + 1

    deduction_start_col = accrued_col + 1
    deduction_end_col = (
        deduction_start_col + len(used_deduction_type_ids) - 1
        if used_deduction_type_ids
        else deduction_start_col - 1
    )
    deduction_total_col = deduction_end_col + 1
    total_to_pay_col = deduction_total_col + 1

    subdivision_totals: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))

    for row in rows:
        payment_format_obj = getattr(row.position, "payment_format", None) if row.position else None
        payment_format = getattr(payment_format_obj, "name", None)
        payment_mode = getattr(payment_format_obj, "calculation_mode", None)
        fired = bool(getattr(row.user, "fired", False))
        base_to_pay = _quantize(row.total_to_pay)
        row_to_pay = base_to_pay
        override = overrides.get(row.user.id) if overrides else None
        if isinstance(override, dict):
            recalc_needed = False
            if "fact_hours" in override:
                row.fact_hours = _decimal(override.get("fact_hours"))
                recalc_needed = True
            if "night_hours" in override:
                row.night_hours = _decimal(override.get("night_hours"))
                recalc_needed = True
            if "rate" in override:
                row.rate = _decimal(override.get("rate"))
                recalc_needed = True
            if "accrual_amount" in override:
                row.total_accruals = _decimal(override.get("accrual_amount"))
            if "deduction_amount" in override:
                row.total_deductions = _decimal(override.get("deduction_amount"))
            row.total_adjustments = row.total_accruals - row.total_deductions

            if recalc_needed and payment_mode and row.position and not row.adjustments_only:
                hours_per_shift = (
                    _decimal(getattr(row.position, "hours_per_shift", None))
                    if getattr(row.position, "hours_per_shift", None) is not None
                    else None
                )
                monthly_shift_norm = (
                    _decimal(getattr(row.position, "monthly_shift_norm", None))
                    if getattr(row.position, "monthly_shift_norm", None) is not None
                    else None
                )
                night_enabled = bool(getattr(row.position, "night_bonus_enabled", False))
                night_percent = _decimal(getattr(row.position, "night_bonus_percent", None))
                factor_for_calc = salary_factor if payment_mode != "fixed" else Decimal("1")
                base_amount, fact_shifts, night_bonus, hours_total = _calc_amounts(
                    rate=row.rate,
                    payment_mode=payment_mode,
                    fact_hours=row.fact_hours,
                    hours_per_shift=hours_per_shift,
                    monthly_shift_norm=monthly_shift_norm,
                    night_hours=row.night_hours,
                    night_bonus_enabled=night_enabled,
                    night_bonus_percent=night_percent,
                    salary_factor=factor_for_calc,
                )
                row.base_amount = base_amount
                row.fact_shifts = fact_shifts
                row.night_bonus = night_bonus
                row.hours_total = hours_total

            if override.get("final_amount") is not None:
                row_to_pay = _quantize(_decimal(override.get("final_amount")))
            elif override.get("calculated_amount") is not None:
                row_to_pay = _quantize(_decimal(override.get("calculated_amount")))
        elif override is not None:
            row_to_pay = _quantize(_decimal(override))

        line = [
            getattr(row.restaurant, "name", None),
            row.subdivision_name,
            getattr(row.position, "name", None),
            f"{row.user.last_name or ''} {row.user.first_name or ''}".strip() or row.user.username,
            payment_format,
            float(row.rate),
            float(row.fact_hours),
            float(row.fact_shifts) if row.fact_shifts is not None else None,
            float(row.night_hours),
            float(row.base_amount),
            float(row.night_bonus),
            float(row.hours_total),
        ]
        for t_id in used_accrual_type_ids:
            value = row.adjustments_by_type.get(t_id, Decimal("0"))
            line.append(float(value))

        line.append(float(row.total_accruals))
        if statement_kind_code == "salary":
            # Statement mode: align with UI "Итого начислено" = Итог + Удержания.
            total_accrued = _quantize(row_to_pay + _decimal(row.total_deductions))
        elif statement_kind_code == "advance":
            # Statement mode: align with UI "Итого начислено" = Итог.
            total_accrued = _quantize(row_to_pay)
        else:
            total_accrued = _quantize(row.hours_total + row.total_accruals)
            if payment_mode == "fixed":
                total_accrued = _quantize(total_accrued * salary_factor)
            # Manual payout correction should shift "Итого начислено" by the same delta
            # as "К выплате" to keep payroll math consistent with statement UI.
            payout_delta = _quantize(row_to_pay - base_to_pay)
            if payout_delta != Decimal("0"):
                total_accrued = _quantize(total_accrued + payout_delta)
        line.append(float(total_accrued))

        for t_id in used_deduction_type_ids:
            value = row.adjustments_by_type.get(t_id, Decimal("0"))
            line.append(float(-value) if value < 0 else float(value))

        line.extend(
            [
                float(row.total_deductions),
                float(row_to_pay),
            ]
        )
        sub_key = row.subdivision_name or "Без подразделения"
        subtotal = subdivision_totals[sub_key]
        subtotal["base_amount"] += row.base_amount
        subtotal["night_bonus"] += row.night_bonus
        subtotal["hours_total"] += row.hours_total
        subtotal["accruals"] += row.total_accruals
        subtotal["deductions"] += row.total_deductions
        subtotal["total_to_pay"] += row_to_pay
        if key_deduction_type_ids:
            key_totals = subtotal.setdefault("key_deductions", defaultdict(Decimal))
            for type_id, canonical in key_deduction_type_ids.items():
                amount = _decimal(row.adjustments_by_type.get(type_id))
                if amount < 0:
                    amount = -amount
                key_totals[canonical] += amount
        ws.append(line)
        if fired:
            ws.cell(row=ws.max_row, column=employee_col).fill = FIRED_FILL

    data_rows_count = len(rows)
    total_line = [""] * len(all_headers) if rows else None
    if total_line:
        total_line[0] = "Итого"

    data_end_row = 1 + data_rows_count if data_rows_count else 1
    if data_rows_count:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{data_end_row}"
        spacer_row_idx = data_end_row + 1
        ws.append([""] * len(all_headers))
        spacer_start = ws.cell(row=spacer_row_idx, column=1).coordinate
        spacer_end = ws.cell(row=spacer_row_idx, column=ws.max_column).coordinate
        ws.merge_cells(f"{spacer_start}:{spacer_end}")
        spacer_cell = ws.cell(row=spacer_row_idx, column=1)
        spacer_cell.fill = SPACER_FILL
        total_row_idx = spacer_row_idx + 1
    else:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
        total_row_idx = data_end_row + 1

    if total_line:
        ws.append(total_line)
        total_row_idx = ws.max_row
        _apply_payroll_total_formulas(
            ws,
            total_row_idx,
            data_end_row,
            base_amount_col=base_amount_col,
            night_bonus_col=night_bonus_col,
            total_hours_col=total_hours_col,
            accrual_start_col=accrual_start_col,
            accrual_end_col=accrual_end_col,
            accrual_total_col=accrual_total_col,
            accrued_col=accrued_col,
            deduction_start_col=deduction_start_col,
            deduction_end_col=deduction_end_col,
            deduction_total_col=deduction_total_col,
            total_to_pay_col=total_to_pay_col,
        )

    main_currency_cols = {
        headers.index(name) + 1
        for name in {"Ставка", "Сумма за часы", "Надбавка за ночь", "Итого за часы"}
    }
    main_number_cols = {
        headers.index(name) + 1
        for name in {"Часы", "Смены", "Ночные часы"}
    }

    for col_idx in range(accrual_start_col, accrual_end_col + 1):
        main_currency_cols.add(col_idx)
    main_currency_cols.add(accrual_total_col)
    main_currency_cols.add(accrued_col)

    for col_idx in range(deduction_start_col, deduction_end_col + 1):
        main_currency_cols.add(col_idx)
    main_currency_cols.add(deduction_total_col)
    main_currency_cols.add(total_to_pay_col)
    _apply_number_formats(ws, currency_cols=main_currency_cols, number_cols=main_number_cols)
    _style_sheet(ws, freeze="E2")

    if total_line:
        total_cells = next(
            ws.iter_rows(min_row=total_row_idx, max_row=total_row_idx, min_col=1, max_col=ws.max_column)
        )
        for cell in total_cells:
            cell.fill = TOTAL_FILL
        total_to_pay_cell = ws.cell(row=total_row_idx, column=ws.max_column)
        try:
            if total_to_pay_cell.value is not None and float(total_to_pay_cell.value) < 0:
                total_to_pay_cell.fill = NEG_TOTAL_FILL
        except (TypeError, ValueError):
            pass
    _autosize_sheet(ws)
    return subdivision_totals


def _build_subdivision_sheet(
    wb: Workbook,
    *,
    subdivision_totals: Mapping[str, Mapping[str, Decimal]],
    revenue_amount: Optional[Decimal] = None,
    dashboard_title: Optional[str] = None,
    sheet_name: str = "Подразделения",
) -> None:
    ws_sub = wb.create_sheet(sheet_name)
    total_to_pay_all = sum(
        (_decimal(agg.get("total_to_pay", Decimal("0"))) for agg in subdivision_totals.values()),
        Decimal("0"),
    )
    headers = [
        "Подразделение",
        "Сумма за часы",
        "Надбавка за ночь",
        "Итого за часы",
        "Начисления",
        *KEY_DEDUCTION_NAMES,
        "Удержания",
        "К выплате",
        "Доля ФОТ",
    ]
    ws_sub.append(headers)
    for sub_name in sorted(subdivision_totals.keys(), key=lambda v: (v or "").lower()):
        agg = subdivision_totals[sub_name]
        key_totals = agg.get("key_deductions", {})
        total_to_pay = _decimal(agg.get("total_to_pay", Decimal("0")))
        pay_share = (total_to_pay / total_to_pay_all) if total_to_pay_all > 0 else None
        ws_sub.append(
            [
                sub_name,
                float(agg.get("base_amount", Decimal("0"))),
                float(agg.get("night_bonus", Decimal("0"))),
                float(agg.get("hours_total", Decimal("0"))),
                float(agg.get("accruals", Decimal("0"))),
                *[float(_decimal(key_totals.get(name))) for name in KEY_DEDUCTION_NAMES],
                float(agg.get("deductions", Decimal("0"))),
                float(total_to_pay),
                float(pay_share) if pay_share is not None else None,
            ]
        )

    data_start_row = 2
    data_end_row = ws_sub.max_row

    sub_currency_cols = set(range(2, 10))
    _apply_number_formats(ws_sub, currency_cols=sub_currency_cols, number_cols=set())
    if data_end_row >= data_start_row:
        for row_idx in range(data_start_row, data_end_row + 1):
            ws_sub.cell(row=row_idx, column=10).number_format = NUMBER_FORMAT_PERCENT
    if ws_sub.max_row:
        ws_sub.auto_filter.ref = f"A1:{get_column_letter(ws_sub.max_column)}{ws_sub.max_row}"
    _style_sheet(ws_sub)

    summary_row = ws_sub.max_row + 2
    ws_sub.cell(row=summary_row, column=1, value="Итого ФОТ (к выплате)")
    total_cell = ws_sub.cell(row=summary_row, column=2, value=float(total_to_pay_all))
    total_cell.number_format = NUMBER_FORMAT_CURRENCY

    if revenue_amount is not None:
        safe_revenue = _decimal(revenue_amount)
        ws_sub.cell(row=summary_row + 1, column=1, value="Выручка ресторана")
        revenue_cell = ws_sub.cell(row=summary_row + 1, column=2, value=float(safe_revenue))
        revenue_cell.number_format = NUMBER_FORMAT_CURRENCY

        ws_sub.cell(row=summary_row + 2, column=1, value="ФОТ / Выручка")
        if safe_revenue > 0:
            ratio_cell = ws_sub.cell(row=summary_row + 2, column=2, value=float(total_to_pay_all / safe_revenue))
            ratio_cell.number_format = NUMBER_FORMAT_PERCENT
        else:
            ws_sub.cell(row=summary_row + 2, column=2, value="—")

    for row_idx in range(summary_row, summary_row + (3 if revenue_amount is not None else 1)):
        ws_sub.cell(row=row_idx, column=1).font = SUBDIVISION_TOTAL_FONT
        ws_sub.cell(row=row_idx, column=1).fill = SUBDIVISION_TOTAL_FILL
        ws_sub.cell(row=row_idx, column=2).fill = SUBDIVISION_TOTAL_FILL
        ws_sub.cell(row=row_idx, column=1).border = THIN_BORDER
        ws_sub.cell(row=row_idx, column=2).border = THIN_BORDER

    _autosize_sheet(ws_sub)


def _build_consolidated_subdivision_sheet(
    wb: Workbook,
    *,
    groups: Sequence[tuple[str, Mapping[str, Mapping[str, Decimal]]]],
    sheet_name: str = "Подразделения",
) -> None:
    ws_sub = wb.create_sheet(sheet_name)
    headers = [
        "Подразделение",
        "Сумма за часы",
        "Надбавка за ночь",
        "Итого за часы",
        "Начисления",
        *KEY_DEDUCTION_NAMES,
        "Удержания",
        "К выплате",
    ]
    ws_sub.append(headers)

    spacer_rows: list[int] = []
    restaurant_rows: list[int] = []
    total_rows: list[int] = []
    grand_total_row_idx: Optional[int] = None
    grand_base = Decimal("0")
    grand_night = Decimal("0")
    grand_hours = Decimal("0")
    grand_accruals = Decimal("0")
    grand_deductions = Decimal("0")
    grand_pay = Decimal("0")
    grand_key_totals: dict[str, Decimal] = defaultdict(Decimal)

    for idx, (restaurant_name, totals) in enumerate(groups):
        if idx:
            spacer_row_idx = ws_sub.max_row + 1
            ws_sub.append([""] * len(headers))
            spacer_rows.append(spacer_row_idx)

        rest_row_idx = ws_sub.max_row + 1
        ws_sub.append([restaurant_name] + [""] * (len(headers) - 1))
        ws_sub.merge_cells(
            f"A{rest_row_idx}:{get_column_letter(len(headers))}{rest_row_idx}"
        )
        restaurant_rows.append(rest_row_idx)

        ordered_subs = sorted(totals.keys(), key=lambda v: (v or "").lower())
        for sub_name in ordered_subs:
            agg = totals[sub_name]
            key_totals = agg.get("key_deductions", {})
            ws_sub.append(
                [
                    sub_name,
                    float(agg.get("base_amount", Decimal("0"))),
                    float(agg.get("night_bonus", Decimal("0"))),
                    float(agg.get("hours_total", Decimal("0"))),
                    float(agg.get("accruals", Decimal("0"))),
                    *[float(_decimal(key_totals.get(name))) for name in KEY_DEDUCTION_NAMES],
                    float(agg.get("deductions", Decimal("0"))),
                    float(agg.get("total_to_pay", Decimal("0"))),
                ]
            )

        base_total = sum((totals[sub].get("base_amount", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        night_total = sum((totals[sub].get("night_bonus", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        hours_total = sum((totals[sub].get("hours_total", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        accruals_total = sum((totals[sub].get("accruals", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        deductions_total = sum((totals[sub].get("deductions", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        pay_total = sum((totals[sub].get("total_to_pay", Decimal("0")) for sub in ordered_subs), Decimal("0"))
        key_totals_sum = {
            name: sum(
                (_decimal(totals[sub].get("key_deductions", {}).get(name)) for sub in ordered_subs),
                Decimal("0"),
            )
            for name in KEY_DEDUCTION_NAMES
        }
        grand_base += base_total
        grand_night += night_total
        grand_hours += hours_total
        grand_accruals += accruals_total
        grand_deductions += deductions_total
        grand_pay += pay_total
        for name, value in key_totals_sum.items():
            grand_key_totals[name] += value

        if ordered_subs:
            total_row_idx = ws_sub.max_row + 1
            ws_sub.append(
                [
                    "Итого",
                    float(base_total),
                    float(night_total),
                    float(hours_total),
                    float(accruals_total),
                    *[float(key_totals_sum[name]) for name in KEY_DEDUCTION_NAMES],
                    float(deductions_total),
                    float(pay_total),
                ]
            )
            total_rows.append(total_row_idx)

    if groups:
        grand_total_row_idx = ws_sub.max_row + 1
        ws_sub.append(
            [
                "Итого (все рестораны)",
                float(grand_base),
                float(grand_night),
                float(grand_hours),
                float(grand_accruals),
                *[float(grand_key_totals.get(name, Decimal("0"))) for name in KEY_DEDUCTION_NAMES],
                float(grand_deductions),
                float(grand_pay),
            ]
        )

    sub_currency_cols = set(range(2, len(headers) + 1))
    _apply_number_formats(ws_sub, currency_cols=sub_currency_cols, number_cols=set())
    if ws_sub.max_row:
        ws_sub.auto_filter.ref = f"A1:{get_column_letter(ws_sub.max_column)}{ws_sub.max_row}"
    _style_sheet(ws_sub)

    for row_idx in spacer_rows:
        start = ws_sub.cell(row=row_idx, column=1).coordinate
        end = ws_sub.cell(row=row_idx, column=len(headers)).coordinate
        ws_sub.merge_cells(f"{start}:{end}")
        ws_sub.cell(row=row_idx, column=1).fill = SPACER_FILL
    for row_idx in restaurant_rows:
        ws_sub.cell(row=row_idx, column=1).fill = SPACER_FILL
    for row_idx in total_rows:
        total_cells = next(
            ws_sub.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=len(headers))
        )
        for cell in total_cells:
            cell.fill = SUBDIVISION_TOTAL_FILL
            cell.font = SUBDIVISION_TOTAL_FONT

    if grand_total_row_idx:
        total_cells = next(
            ws_sub.iter_rows(
                min_row=grand_total_row_idx,
                max_row=grand_total_row_idx,
                min_col=1,
                max_col=len(headers),
            )
        )
        for cell in total_cells:
            cell.fill = SUBDIVISION_TOTAL_FILL
            cell.font = SUBDIVISION_TOTAL_FONT

    _autosize_sheet(ws_sub)


def _build_adjustment_sheets(
    wb: Workbook,
    *,
    adj_rows: Iterable[object],
    adj_type_map: Mapping[int, PayrollAdjustmentType],
    user_map: Mapping[int, User],
    include_restaurant: bool = False,
    include_subdivision: bool = False,
    restaurant_lookup: Optional[Mapping[int, str]] = None,
) -> None:
    ws_adj_accruals = wb.create_sheet("Начисления")
    ws_adj_deductions = wb.create_sheet("Удержания")

    headers: list[str] = []
    if include_restaurant:
        headers.append("Ресторан")
    if include_subdivision:
        headers.append("Подразделение")
    headers.extend(["Сотрудник", "Тип", "Сумма", "Дата", "Комментарий"])

    for ws_adj in (ws_adj_accruals, ws_adj_deductions):
        ws_adj.append(headers)

    rows = list(adj_rows or [])

    def _restaurant_name_for(adj: object, user: Optional[User]) -> Optional[str]:
        name = getattr(adj, "restaurant_name", None)
        if name:
            return name
        rid = getattr(adj, "restaurant_id", None)
        if rid and restaurant_lookup:
            lookup_name = restaurant_lookup.get(rid)
            if lookup_name:
                return lookup_name
        if user and getattr(user, "workplace_restaurant", None):
            return getattr(user.workplace_restaurant, "name", None)
        return None

    def _subdivision_name_for(adj: object, user: Optional[User]) -> Optional[str]:
        name = getattr(adj, "subdivision_name", None)
        if name:
            return name
        if user and getattr(user, "position", None) and getattr(user.position, "restaurant_subdivision", None):
            return getattr(user.position.restaurant_subdivision, "name", None)
        return None

    def _sort_key(adj: object) -> tuple[str, str, str, str]:
        user = user_map.get(getattr(adj, "user_id", None))
        restaurant_name = (_restaurant_name_for(adj, user) or "").lower()
        last = (getattr(user, "last_name", None) or "").lower() if user else ""
        first = (getattr(user, "first_name", None) or "").lower() if user else ""
        adj_type = adj_type_map.get(getattr(adj, "adjustment_type_id", None))
        type_name = (getattr(adj_type, "name", None) or "").lower()
        return (restaurant_name, last, first, type_name)

    rows.sort(key=_sort_key)

    for adj in rows:
        adj_type = adj_type_map.get(getattr(adj, "adjustment_type_id", None))
        user = user_map.get(getattr(adj, "user_id", None))
        if not adj_type or not user:
            continue
        amount = _decimal(getattr(adj, "amount", None))
        if adj_type.kind == "deduction":
            amount = amount * Decimal("-1")
            target_ws = ws_adj_deductions
        else:
            target_ws = ws_adj_accruals

        row_values: list[object] = []
        if include_restaurant:
            row_values.append(_restaurant_name_for(adj, user))
        if include_subdivision:
            row_values.append(_subdivision_name_for(adj, user))

        row_values.extend(
            [
                f"{user.last_name or ''} {user.first_name or ''}".strip() or user.username,
                adj_type.name,
                float(amount),
                adj.date.isoformat() if getattr(adj, "date", None) else None,
                getattr(adj, "comment", None),
            ]
        )
        target_ws.append(row_values)

    pre_cols = int(include_restaurant) + int(include_subdivision)
    amount_col = pre_cols + 3
    date_col = pre_cols + 4
    adj_currency_cols = {amount_col}
    adj_date_cols = {date_col}
    _apply_number_formats(
        ws_adj_accruals,
        currency_cols=adj_currency_cols,
        number_cols=set(),
        date_cols=adj_date_cols,
    )
    _apply_number_formats(
        ws_adj_deductions,
        currency_cols=adj_currency_cols,
        number_cols=set(),
        date_cols=adj_date_cols,
    )
    if ws_adj_accruals.max_row:
        ws_adj_accruals.auto_filter.ref = (
            f"A1:{get_column_letter(ws_adj_accruals.max_column)}{ws_adj_accruals.max_row}"
        )
    if ws_adj_deductions.max_row:
        ws_adj_deductions.auto_filter.ref = (
            f"A1:{get_column_letter(ws_adj_deductions.max_column)}{ws_adj_deductions.max_row}"
        )
    _style_sheet(ws_adj_accruals)
    _style_sheet(ws_adj_deductions)
    _autosize_sheet(ws_adj_accruals)
    _autosize_sheet(ws_adj_deductions)


def build_payroll_report(
    db: Session,
    *,
    period: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    company_id: Optional[int],
    restaurant_id: Optional[int],
    subdivision_id: Optional[int],
    user_ids: Optional[Sequence[int]] = None,
    salary_percent: Optional[float] = None,
    overrides: Optional[dict[int, object]] = None,
    snapshot: Optional[PayrollReportSnapshot] = None,
) -> BytesIO:
    salary_factor = (
        Decimal(str(salary_percent)) / Decimal("100") if salary_percent is not None else Decimal("1")
    )
    report_start: Optional[date] = None
    report_end: Optional[date] = None
    effective_company_id = company_id
    effective_restaurant_id = restaurant_id

    if snapshot:
        results = snapshot.results
        adj_type_map = snapshot.adj_type_map
        adj_rows = snapshot.adj_rows
        user_map = snapshot.user_map
        salary_factor = snapshot.salary_factor
        report_start = snapshot.date_from
        report_end = snapshot.date_to
        if effective_company_id is None:
            effective_company_id = snapshot.company_id
        if effective_restaurant_id is None:
            effective_restaurant_id = snapshot.restaurant_id
    else:
        start, end = resolve_payroll_export_bounds(period=period, date_from=date_from, date_to=date_to)
        report_start = start
        report_end = end

        requested_user_ids: Optional[list[int]] = None
        if user_ids is not None:
            requested_user_ids = list({int(uid) for uid in user_ids})
            if not requested_user_ids:
                return BytesIO()
    
        attendance = _collect_attendance(
            db,
            start,
            end,
            company_id=company_id,
            restaurant_id=restaurant_id,
            subdivision_id=subdivision_id,
            user_ids=requested_user_ids,
        )
        attendance_user_ids = set(attendance.keys())
    
        # adjustments (включаем пользователей только с корректировками)
        adj_by_user, adj_type_map, adj_rows = _collect_adjustments(
            db,
            start,
            end,
            requested_user_ids,
            company_id=company_id,
            restaurant_id=restaurant_id,
            subdivision_id=subdivision_id,
        )
        adj_restaurant_ids: dict[int, set[int]] = defaultdict(set)
        for adj in adj_rows:
            if getattr(adj, "restaurant_id", None) is not None:
                adj_restaurant_ids[adj.user_id].add(adj.restaurant_id)
    
        # Fixed-salary users should be included even без смен/корректировок.
        fixed_user_ids: set[int] = set()
        fixed_q = (
            db.query(User.id)
            .join(Position, Position.id == User.position_id)
            .join(PaymentFormat, PaymentFormat.id == Position.payment_format_id)
            .filter(PaymentFormat.calculation_mode == "fixed")
        )
        if company_id:
            fixed_q = fixed_q.filter(User.company_id == company_id)
        if restaurant_id:
            fixed_q = fixed_q.filter(User.workplace_restaurant_id == restaurant_id)
        if subdivision_id:
            fixed_q = fixed_q.filter(Position.restaurant_subdivision_id == subdivision_id)
        if requested_user_ids is not None:
            fixed_q = fixed_q.filter(User.id.in_(requested_user_ids))
        fixed_user_ids = {row[0] for row in fixed_q.distinct().all()}
    
        # include attendance + adjustments users + fixed salary users by default
        candidate_user_ids = set(attendance_user_ids) | set(adj_by_user.keys()) | fixed_user_ids
        if requested_user_ids:
            candidate_user_ids.update(requested_user_ids)
        if not candidate_user_ids:
            return BytesIO()
    
        for uid in candidate_user_ids:
            attendance.setdefault(uid, [])
    
        users = (
            db.query(User)
            .options(
                joinedload(User.position).joinedload(Position.payment_format),
                joinedload(User.position).joinedload(Position.restaurant_subdivision),
                joinedload(User.company),
                joinedload(User.workplace_restaurant),
                joinedload(User.restaurants),
            )
            .filter(User.id.in_(candidate_user_ids) if candidate_user_ids else False)  # short-circuit to return empty set
            .all()
        )
        user_map = {u.id: u for u in users}
    
        restaurant_ids: set[int] = set()
        if restaurant_id:
            restaurant_ids.add(restaurant_id)
        for stats_list in attendance.values():
            for stats in stats_list:
                rid = stats.get("restaurant_id")
                if rid:
                    restaurant_ids.add(rid)
        for user in users:
            rid = getattr(user, "workplace_restaurant_id", None)
            if rid:
                restaurant_ids.add(rid)
        restaurant_map = {}
        if restaurant_ids:
            restaurant_map = {
                r.id: r for r in db.query(Restaurant).filter(Restaurant.id.in_(restaurant_ids)).all()
            }
    
        position_ids: set[int] = set()
        for stats_list in attendance.values():
            for stats in stats_list:
                if stats.get("position_id"):
                    position_ids.add(stats["position_id"])
        position_map = {}
        if position_ids:
            position_map = {
                p.id: p
                for p in db.query(Position)
                .options(joinedload(Position.payment_format), joinedload(Position.restaurant_subdivision))
                .filter(Position.id.in_(position_ids))
                .all()
            }
    
        # Build per-user rows
        results: list[UserPayrollRow] = []
    
        requested_user_id_set = set(requested_user_ids or [])
    
        def _passes_filters(user: User, stats: dict) -> bool:
            if company_id and getattr(user, "company_id", None) != company_id:
                return False
            if restaurant_id:
                has_attendance_in_restaurant = stats.get("restaurant_id") == restaurant_id
                has_workplace_in_restaurant = getattr(user, "workplace_restaurant_id", None) == restaurant_id
                if not has_attendance_in_restaurant and not has_workplace_in_restaurant and user.id not in requested_user_id_set:
                    return False
            if subdivision_id:
                pos = getattr(user, "position", None)
                if getattr(pos, "restaurant_subdivision_id", None) != subdivision_id:
                    return False
            return True
    
        for user_id in candidate_user_ids:
            user = user_map.get(user_id)
            if not user:
                continue
    
            stats_list = attendance.get(user_id, [])
            adj_map_full = adj_by_user.get(user_id, {})
            requested = user_id in requested_user_id_set
    
            is_fixed_user = user_id in fixed_user_ids
    
            has_adjustments_in_selected_restaurant = False
            if restaurant_id and adj_map_full:
                has_adjustments_in_selected_restaurant = restaurant_id in adj_restaurant_ids.get(user_id, set())
                if has_adjustments_in_selected_restaurant:
                    has_stats_in_restaurant = any(
                        stat.get("restaurant_id") == restaurant_id for stat in stats_list
                    )
                    if not has_stats_in_restaurant:
                        stats_list = list(stats_list) if stats_list else []
                        stats_list.append(
                            {
                                "position_id": getattr(user, "position_id", None),
                                "rate": None,
                                "duration_minutes": 0,
                                "night_minutes": 0,
                                "restaurant_id": restaurant_id,
                                "adjustments_only": True,
                            }
                        )
    
            if not stats_list and not adj_map_full and not requested and not is_fixed_user:
                continue
            if not stats_list:
                fallback_restaurant = getattr(user, "workplace_restaurant_id", None)
                if fallback_restaurant is None and user.restaurants:
                    first_restaurant = next((r for r in user.restaurants if getattr(r, "id", None) is not None), None)
                    if first_restaurant:
                        fallback_restaurant = first_restaurant.id
                stats_list = [
                    {
                        "position_id": getattr(user, "position_id", None),
                        "rate": None,
                        "duration_minutes": 0,
                        "night_minutes": 0,
                        "restaurant_id": fallback_restaurant,
                    }
                ]
    
            main_position_id = getattr(user, "position_id", None)
            adjust_position_id = None
            if main_position_id is not None and any(
                stats.get("position_id") == main_position_id for stats in stats_list
            ):
                adjust_position_id = main_position_id
            elif stats_list:
                adjust_position_id = stats_list[0].get("position_id")
            adjustments_applied = False
    
            for stats in stats_list:
                if not _passes_filters(user, stats):
                    continue
    
                pos_id = stats.get("position_id")
                position = None
                if pos_id:
                    position = position_map.get(pos_id) or (getattr(user, "position", None) if pos_id == main_position_id else None)
                else:
                    position = getattr(user, "position", None)
    
                company = getattr(user, "company", None)
                restaurant_obj = None
                # pick restaurant by filter or from attendance
                if restaurant_id:
                    restaurant_obj = restaurant_map.get(restaurant_id)
                if restaurant_obj is None:
                    if stats.get("restaurant_id"):
                        rid = stats["restaurant_id"]
                        restaurant_obj = restaurant_map.get(rid)
                if restaurant_obj is None:
                    workplace_id = getattr(user, "workplace_restaurant_id", None)
                    if workplace_id:
                        restaurant_obj = restaurant_map.get(workplace_id) or getattr(user, "workplace_restaurant", None)
                if restaurant_obj is None and user.restaurants:
                    restaurant_obj = user.restaurants[0]
    
                minutes = stats.get("duration_minutes", 0) or 0
                night_minutes = stats.get("night_minutes", 0) or 0
                adjustments_only = bool(stats.get("adjustments_only"))
                stats_restaurant_id = stats.get("restaurant_id")
                workplace_restaurant_id = getattr(user, "workplace_restaurant_id", None)
                is_workplace_row = (
                    workplace_restaurant_id is not None
                    and stats_restaurant_id == workplace_restaurant_id
                )

                payment_format = getattr(position, "payment_format", None) if position else None
                payment_mode = getattr(payment_format, "calculation_mode", None)

                # For restaurant-scoped exports, fixed-salary base applies only to the
                # workplace restaurant; other restaurants should show adjustments only.
                if (
                    payment_mode == "fixed"
                    and restaurant_id
                    and workplace_restaurant_id
                    and workplace_restaurant_id != restaurant_id
                ):
                    adjustments_only = True

                # Fixed-salary base should only appear in the workplace restaurant row.
                if adjustments_only and payment_mode == "fixed" and is_workplace_row:
                    adjustments_only = False

                fact_shifts: Optional[Decimal] = None
                if adjustments_only:
                    fact_hours = Decimal("0")
                    night_hours = Decimal("0")
                    rate = Decimal("0")
                    base_amount = Decimal("0")
                    night_bonus = Decimal("0")
                    hours_total = Decimal("0")
                else:
                    fact_hours = Decimal(minutes) / Decimal(60) if minutes else Decimal("0")
                    night_hours = Decimal(night_minutes) / Decimal(60) if night_minutes else Decimal("0")
    
                    hours_per_shift = (
                        _decimal(getattr(position, "hours_per_shift", None))
                        if position and getattr(position, "hours_per_shift", None)
                        else None
                    )
                    monthly_shift_norm = (
                        _decimal(getattr(position, "monthly_shift_norm", None))
                        if position and getattr(position, "monthly_shift_norm", None)
                        else None
                    )
                    night_enabled = bool(getattr(position, "night_bonus_enabled", False)) if position else False
                    night_percent = _decimal(getattr(position, "night_bonus_percent", None)) if position else Decimal("0")
                    rate = _resolve_rate(
                        user,
                        position,
                        payment_mode,
                        stats_rate=stats.get("rate"),
                    )
    
                    factor_for_calc = salary_factor if payment_mode != "fixed" else Decimal("1")
    
                    base_amount, fact_shifts, night_bonus, hours_total = _calc_amounts(
                        rate=rate,
                        payment_mode=payment_mode,
                        fact_hours=fact_hours,
                        hours_per_shift=hours_per_shift,
                        monthly_shift_norm=monthly_shift_norm,
                        night_hours=night_hours,
                        night_bonus_enabled=night_enabled,
                        night_bonus_percent=night_percent,
                        salary_factor=factor_for_calc,
                    )
    
                adj_map = {}
                if not adjustments_applied and (pos_id == adjust_position_id or adjust_position_id is None):
                    adj_map = adj_map_full
                    adjustments_applied = True
    
                total_adjustments = Decimal("0")
                total_accruals = Decimal("0")
                total_deductions = Decimal("0")
                for type_id, value in adj_map.items():
                    total_adjustments += value
                    if value >= 0:
                        total_accruals += value
                    else:
                        total_deductions += -value
                if payment_mode == "fixed":
                    accrued_for_payout = _quantize((hours_total + total_accruals) * salary_factor)
                    total_to_pay = _quantize(accrued_for_payout - total_deductions)
                else:
                    accrued_for_payout = _quantize(hours_total + total_accruals)
                    total_to_pay = _quantize(hours_total + total_adjustments)
    
                subdivision_name = getattr(position.restaurant_subdivision, "name", None) if position else None
                results.append(
                    UserPayrollRow(
                        user=user,
                        company=company,
                        restaurant=restaurant_obj,
                        subdivision_name=subdivision_name,
                        position=position,
                        rate=rate,
                        fact_hours=fact_hours,
                        fact_shifts=fact_shifts,
                        night_hours=night_hours,
                        base_amount=base_amount,
                        night_bonus=night_bonus,
                        hours_total=hours_total,
                        adjustments_by_type=adj_map,
                        total_adjustments=total_adjustments,
                        total_accruals=total_accruals,
                        total_deductions=total_deductions,
                        total_to_pay=total_to_pay,
                        adjustments_only=adjustments_only,
                    )
                )
    
    report_restaurant_ids: set[int] = set()
    if effective_restaurant_id:
        report_restaurant_ids.add(int(effective_restaurant_id))
    for row in results:
        rid = getattr(getattr(row, "restaurant", None), "id", None)
        if rid is not None:
            try:
                parsed = int(rid)
            except (TypeError, ValueError):
                continue
            if parsed > 0:
                report_restaurant_ids.add(parsed)

    if effective_company_id is None:
        company_ids: set[int] = set()
        for row in results:
            company_id_value = getattr(getattr(row, "company", None), "id", None)
            if company_id_value is None:
                company_id_value = getattr(row.user, "company_id", None)
            if company_id_value is None:
                continue
            try:
                parsed_company_id = int(company_id_value)
            except (TypeError, ValueError):
                continue
            if parsed_company_id > 0:
                company_ids.add(parsed_company_id)
        if len(company_ids) == 1:
            effective_company_id = next(iter(company_ids))

    revenue_amount = _resolve_revenue_amount_for_report(
        db,
        company_id=effective_company_id,
        restaurant_ids=sorted(report_restaurant_ids),
        date_from=report_start,
        date_to=report_end,
    )

    dashboard_title = "ФОТ по подразделениям"
    if effective_restaurant_id and results:
        first_restaurant_name = next(
            (
                getattr(getattr(item, "restaurant", None), "name", None)
                for item in results
                if getattr(getattr(item, "restaurant", None), "name", None)
            ),
            None,
        )
        if first_restaurant_name:
            dashboard_title = f"ФОТ по подразделениям: {first_restaurant_name}"

    wb = Workbook()
    ws = wb.active
    ws.title = "\u0412\u0435\u0434\u043e\u043c\u043e\u0441\u0442\u044c"

    subdivision_totals = _build_payroll_sheet(
        ws,
        results=results,
        adj_type_map=adj_type_map,
        salary_factor=salary_factor,
        overrides=overrides,
        statement_kind=(snapshot.statement_kind if snapshot else None),
    )
    _build_subdivision_sheet(
        wb,
        subdivision_totals=subdivision_totals,
        revenue_amount=revenue_amount,
        dashboard_title=dashboard_title,
    )
    _build_adjustment_sheets(
        wb,
        adj_rows=adj_rows,
        adj_type_map=adj_type_map,
        user_map=user_map,
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _build_payroll_snapshot_from_statement(
    db: Session,
    statement: PayrollAdvanceStatement,
) -> Optional[PayrollReportSnapshot]:
    items = list(getattr(statement, "items", None) or [])
    if not items:
        return None

    salary_factor = (
        Decimal(str(statement.salary_percent)) / Decimal("100")
        if statement.salary_percent is not None
        else Decimal("1")
    )

    user_ids = sorted({item.user_id for item in items if item.user_id is not None})
    if not user_ids:
        return None

    users = (
        db.query(User)
        .options(
            joinedload(User.company),
            joinedload(User.position).joinedload(Position.restaurant_subdivision),
            joinedload(User.workplace_restaurant),
            joinedload(User.restaurants),
        )
        .filter(User.id.in_(user_ids))
        .all()
    )
    user_map = {u.id: u for u in users}

    restaurant_ids: set[int] = set()
    if statement.restaurant_id:
        restaurant_ids.add(statement.restaurant_id)
    for item in items:
        if item.restaurant_id:
            restaurant_ids.add(item.restaurant_id)
    for user in users:
        rid = getattr(user, "workplace_restaurant_id", None)
        if rid:
            restaurant_ids.add(rid)
        for rest in getattr(user, "restaurants", None) or []:
            if rest and getattr(rest, "id", None) is not None:
                restaurant_ids.add(rest.id)
    restaurant_map: dict[int, Restaurant] = {}
    if restaurant_ids:
        restaurant_map = {
            r.id: r for r in db.query(Restaurant).filter(Restaurant.id.in_(restaurant_ids)).all()
        }

    position_ids: set[int] = {item.position_id for item in items if item.position_id}
    position_map: dict[int, Position] = {}
    if position_ids:
        position_map = {
            p.id: p
            for p in db.query(Position)
            .options(joinedload(Position.payment_format), joinedload(Position.restaurant_subdivision))
            .filter(Position.id.in_(position_ids))
            .all()
        }

    adjustments_snapshot = statement.adjustments_snapshot or []
    adj_rows: list[AdjustmentSnapshotRow] = []
    adj_type_map: dict[int, PayrollAdjustmentType] = {}
    adj_by_user_snapshot: dict[int, dict[int, Decimal]] = defaultdict(lambda: defaultdict(Decimal))

    for entry in adjustments_snapshot:
        if not isinstance(entry, dict):
            continue
        user_id = entry.get("user_id")
        type_id = entry.get("adjustment_type_id")
        try:
            user_id = int(user_id) if user_id is not None else None
            type_id = int(type_id) if type_id is not None else None
        except (TypeError, ValueError):
            continue
        if user_id is None or type_id is None:
            continue
        name = entry.get("type_name")
        kind = entry.get("kind")
        if type_id not in adj_type_map and name and kind:
            adj_type_map[type_id] = PayrollAdjustmentType(
                id=type_id,
                name=name,
                kind=kind,
                show_in_report=False,
            )

        amount = _decimal(entry.get("amount"))
        signed = amount
        if kind == "deduction" and signed > 0:
            signed = signed * Decimal("-1")
        adj_by_user_snapshot[user_id][type_id] += signed

        date_raw = entry.get("date")
        date_value = None
        if isinstance(date_raw, date):
            date_value = date_raw
        elif date_raw:
            try:
                date_value = date.fromisoformat(str(date_raw))
            except ValueError:
                date_value = None
        adj_rows.append(
            AdjustmentSnapshotRow(
                user_id=user_id,
                adjustment_type_id=type_id,
                amount=amount,
                date=date_value,
                comment=entry.get("comment"),
            )
        )

    calc_type_ids: set[int] = set()
    results: list[UserPayrollRow] = []

    for item in items:
        user = user_map.get(item.user_id)
        if not user:
            continue

        position = position_map.get(item.position_id) if item.position_id else None
        position_name = item.position_name or (getattr(position, "name", None) if position else None)

        calc_snapshot = item.calc_snapshot if isinstance(item.calc_snapshot, dict) else {}

        payment_mode = calc_snapshot.get("payment_mode")
        payment_format_name = calc_snapshot.get("payment_format_name")
        payment_format = getattr(position, "payment_format", None) if position else None
        if payment_mode is None:
            payment_mode = getattr(payment_format, "calculation_mode", None)
        if payment_format_name is None:
            payment_format_name = getattr(payment_format, "name", None)

        payment_format_obj = None
        if payment_format_name or payment_mode:
            payment_format_obj = SimpleNamespace(
                name=payment_format_name,
                calculation_mode=payment_mode,
            )

        position_obj = None
        if position_name or payment_format_obj or payment_format:
            position_obj = SimpleNamespace(
                name=position_name,
                payment_format=payment_format_obj or payment_format,
            )

        base_present = "base_amount" in calc_snapshot
        night_present = "night_bonus" in calc_snapshot
        hours_present = "hours_total" in calc_snapshot
        shifts_present = (
            "fact_shifts" in calc_snapshot and calc_snapshot.get("fact_shifts") is not None
        )

        base_amount = _decimal(calc_snapshot.get("base_amount")) if base_present else Decimal("0")
        night_bonus = _decimal(calc_snapshot.get("night_bonus")) if night_present else Decimal("0")
        hours_total = _decimal(calc_snapshot.get("hours_total")) if hours_present else Decimal("0")
        fact_shifts = (
            _decimal(calc_snapshot.get("fact_shifts"))
            if shifts_present and calc_snapshot.get("fact_shifts") is not None
            else None
        )

        fact_hours = _decimal(getattr(item, "fact_hours", None))
        night_hours = _decimal(getattr(item, "night_hours", None))
        rate = _decimal(getattr(item, "rate", None))

        needs_calc = (
            not base_present or not night_present or not hours_present or not shifts_present
        )
        if needs_calc and position and payment_mode:
            hours_per_shift = (
                _decimal(getattr(position, "hours_per_shift", None))
                if getattr(position, "hours_per_shift", None) is not None
                else None
            )
            monthly_shift_norm = (
                _decimal(getattr(position, "monthly_shift_norm", None))
                if getattr(position, "monthly_shift_norm", None) is not None
                else None
            )
            night_enabled = bool(getattr(position, "night_bonus_enabled", False))
            night_percent = _decimal(getattr(position, "night_bonus_percent", None))

            factor_for_calc = salary_factor if payment_mode != "fixed" else Decimal("1")
            calc_base, calc_shifts, calc_night, calc_hours = _calc_amounts(
                rate=rate,
                payment_mode=payment_mode,
                fact_hours=fact_hours,
                hours_per_shift=hours_per_shift,
                monthly_shift_norm=monthly_shift_norm,
                night_hours=night_hours,
                night_bonus_enabled=night_enabled,
                night_bonus_percent=night_percent,
                salary_factor=factor_for_calc,
            )
            if not base_present:
                base_amount = calc_base
            if not night_present:
                night_bonus = calc_night
            if not hours_present:
                hours_total = calc_hours
            if not shifts_present:
                fact_shifts = calc_shifts
        elif not hours_present and (base_present or night_present):
            hours_total = _quantize(base_amount + night_bonus)

        raw_adj_map = calc_snapshot.get("adjustments_by_type")
        adj_map: dict[int, Decimal] = {}
        if isinstance(raw_adj_map, dict):
            for type_key, value in raw_adj_map.items():
                try:
                    type_id = int(type_key)
                except (TypeError, ValueError):
                    continue
                adj_map[type_id] = _decimal(value)
                calc_type_ids.add(type_id)
        elif item.user_id in adj_by_user_snapshot:
            adj_map = dict(adj_by_user_snapshot.get(item.user_id, {}))

        if adj_map:
            total_adjustments = sum(adj_map.values(), Decimal("0"))
            total_accruals = sum((val for val in adj_map.values() if val >= 0), Decimal("0"))
            total_deductions = sum((-val for val in adj_map.values() if val < 0), Decimal("0"))
        else:
            total_accruals = _decimal(getattr(item, "accrual_amount", None))
            total_deductions = _decimal(getattr(item, "deduction_amount", None))
            total_adjustments = total_accruals - total_deductions

        total_to_pay = _quantize(_decimal(getattr(item, "final_amount", None)))

        subdivision_name = None
        if position and getattr(position, "restaurant_subdivision", None):
            subdivision_name = getattr(position.restaurant_subdivision, "name", None)

        restaurant_obj = None
        if item.restaurant_id:
            restaurant_obj = restaurant_map.get(item.restaurant_id)
        if restaurant_obj is None and statement.restaurant_id:
            restaurant_obj = restaurant_map.get(statement.restaurant_id)
        if restaurant_obj is None:
            workplace_id = getattr(user, "workplace_restaurant_id", None)
            if workplace_id:
                restaurant_obj = restaurant_map.get(workplace_id) or getattr(user, "workplace_restaurant", None)
        if restaurant_obj is None and getattr(user, "restaurants", None):
            restaurant_obj = user.restaurants[0]

        results.append(
            UserPayrollRow(
                user=user,
                company=getattr(user, "company", None),
                restaurant=restaurant_obj,
                subdivision_name=subdivision_name,
                position=position_obj,
                rate=rate,
                fact_hours=fact_hours,
                fact_shifts=fact_shifts,
                night_hours=night_hours,
                base_amount=base_amount,
                night_bonus=night_bonus,
                hours_total=hours_total,
                adjustments_by_type=adj_map,
                total_adjustments=total_adjustments,
                total_accruals=total_accruals,
                total_deductions=total_deductions,
                total_to_pay=total_to_pay,
                adjustments_only=False,
            )
        )

    if calc_type_ids:
        missing = {type_id for type_id in calc_type_ids if type_id not in adj_type_map}
        if missing:
            for adj_type in db.query(PayrollAdjustmentType).filter(PayrollAdjustmentType.id.in_(missing)).all():
                adj_type_map[adj_type.id] = adj_type
    if adj_type_map:
        # Ensure we use DB rows (including show_in_report flag) when available.
        for adj_type in (
            db.query(PayrollAdjustmentType)
            .filter(PayrollAdjustmentType.id.in_(list(adj_type_map.keys())))
            .all()
        ):
            adj_type_map[adj_type.id] = adj_type

    statement_restaurant = getattr(statement, "restaurant", None)
    statement_restaurant_id = getattr(statement_restaurant, "id", None) or statement.restaurant_id
    statement_restaurant_name = getattr(statement_restaurant, "name", None)
    statement_company_id = getattr(statement_restaurant, "company_id", None)
    if statement_restaurant_id and not statement_restaurant_name:
        rest_obj = restaurant_map.get(statement_restaurant_id)
        statement_restaurant_name = getattr(rest_obj, "name", None)
        if statement_company_id is None:
            statement_company_id = getattr(rest_obj, "company_id", None)
    if statement_company_id is None:
        for row in results:
            company_id_value = getattr(getattr(row, "company", None), "id", None)
            if company_id_value is None:
                company_id_value = getattr(row.user, "company_id", None)
            if company_id_value is not None:
                statement_company_id = company_id_value
                break

    user_subdivision_map: dict[int, str] = {}
    for row in results:
        if row.subdivision_name:
            user_subdivision_map[row.user.id] = row.subdivision_name
    for user in users:
        if user.id in user_subdivision_map:
            continue
        pos = getattr(user, "position", None)
        sub = getattr(pos, "restaurant_subdivision", None) if pos else None
        sub_name = getattr(sub, "name", None)
        if sub_name:
            user_subdivision_map[user.id] = sub_name

    for adj in adj_rows:
        if not isinstance(adj, AdjustmentSnapshotRow):
            continue
        if statement_restaurant_id and adj.restaurant_id is None:
            adj.restaurant_id = statement_restaurant_id
        if statement_restaurant_name and not adj.restaurant_name:
            adj.restaurant_name = statement_restaurant_name
        if not adj.subdivision_name:
            adj.subdivision_name = user_subdivision_map.get(adj.user_id)
        adj.statement_id = statement.id

    snapshot = PayrollReportSnapshot(
        results=results,
        adj_type_map=adj_type_map,
        adj_rows=adj_rows,
        user_map=user_map,
        salary_factor=salary_factor,
        date_from=statement.date_from,
        date_to=statement.date_to,
        company_id=statement_company_id,
        restaurant_id=statement_restaurant_id,
        statement_kind=_normalize_statement_kind(getattr(statement, "statement_kind", None)),
    )
    return snapshot


def build_payroll_report_from_statement(
    db: Session,
    statement: PayrollAdvanceStatement,
) -> BytesIO:
    snapshot = _build_payroll_snapshot_from_statement(db, statement)
    if not snapshot:
        return BytesIO()
    return build_payroll_report(
        db,
        company_id=None,
        restaurant_id=None,
        subdivision_id=None,
        snapshot=snapshot,
    )


def build_consolidated_payroll_report_from_statements(
    db: Session,
    statements: Sequence[PayrollAdvanceStatement],
) -> BytesIO:
    statement_list = [s for s in (statements or []) if s is not None]
    if not statement_list:
        return BytesIO()

    reserved_titles = {
        "\u041f\u043e\u0434\u0440\u0430\u0437\u0434\u0435\u043b\u0435\u043d\u0438\u044f",
        "\u041d\u0430\u0447\u0438\u0441\u043b\u0435\u043d\u0438\u044f",
        "\u0423\u0434\u0435\u0440\u0436\u0430\u043d\u0438\u044f",
    }
    used_titles = set(reserved_titles)

    wb = Workbook()
    first_ws = wb.active
    first_ws_used = False

    groups: list[tuple[str, Mapping[str, Mapping[str, Decimal]]]] = []
    combined_adj_rows: list[object] = []
    combined_adj_type_map: dict[int, PayrollAdjustmentType] = {}
    combined_user_map: dict[int, User] = {}
    restaurant_lookup: dict[int, str] = {}

    for statement in statement_list:
        snapshot = _build_payroll_snapshot_from_statement(db, statement)
        if not snapshot:
            continue

        adj_type_map = snapshot.adj_type_map

        restaurant_name = getattr(getattr(statement, "restaurant", None), "name", None)
        base_name = restaurant_name or statement.title or (
            f"Restaurant {statement.restaurant_id}" if statement.restaurant_id else f"Statement {statement.id}"
        )
        sheet_title = _unique_sheet_title(base_name, used_titles)

        ws = first_ws if not first_ws_used else wb.create_sheet(sheet_title)
        ws.title = sheet_title
        first_ws_used = True

        for adj in snapshot.adj_rows:
            if not isinstance(adj, AdjustmentSnapshotRow):
                continue
            adj.restaurant_name = adj.restaurant_name or sheet_title
            adj.restaurant_id = adj.restaurant_id or statement.restaurant_id
            adj.statement_id = adj.statement_id or statement.id
        if statement.restaurant_id:
            restaurant_lookup[statement.restaurant_id] = sheet_title

        sub_totals = _build_payroll_sheet(
            ws,
            results=snapshot.results,
            adj_type_map=adj_type_map,
            salary_factor=snapshot.salary_factor,
            statement_kind=_normalize_statement_kind(getattr(statement, "statement_kind", None)),
        )
        groups.append((sheet_title, sub_totals))

        combined_adj_rows.extend(snapshot.adj_rows)
        combined_adj_type_map.update(snapshot.adj_type_map)
        combined_user_map.update(snapshot.user_map)

    if not first_ws_used:
        return BytesIO()

    _build_consolidated_subdivision_sheet(wb, groups=groups)
    _build_adjustment_sheets(
        wb,
        adj_rows=combined_adj_rows,
        adj_type_map=combined_adj_type_map,
        user_map=combined_user_map,
        include_restaurant=True,
        include_subdivision=True,
        restaurant_lookup=restaurant_lookup,
    )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
