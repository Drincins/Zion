from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment

from . import common as c

router = APIRouter()


@router.get("/timesheet/options", response_model=c.TimesheetOptionsResponse)
def list_timesheet_options(
    db: c.Session = Depends(c.get_db),
    current_user: c.User = Depends(c.get_current_user),
):
    c._ensure_timesheet_export(current_user)
    allowed_workplaces = c._get_allowed_workplace_ids(db, current_user)
    cache_key = c._ids_scope_key(allowed_workplaces)

    def _load_timesheet_options() -> dict:
        restaurants_query = db.query(c.Restaurant).order_by(
            c.func.lower(c.Restaurant.name).nullslast(),
            c.Restaurant.id.asc(),
        )
        if allowed_workplaces is not None:
            if allowed_workplaces:
                restaurants_query = restaurants_query.filter(c.Restaurant.id.in_(allowed_workplaces))
            else:
                restaurants_query = restaurants_query.filter(False)

        restaurants = restaurants_query.all()
        positions_query = (
            db.query(c.Position)
            .options(c.joinedload(c.Position.restaurant_subdivision))
            .order_by(c.Position.name.asc())
        )
        if allowed_workplaces is not None:
            if allowed_workplaces:
                accessible_position_ids = (
                    db.query(c.User.position_id)
                    .filter(
                        c.User.position_id.isnot(None),
                        c.or_(
                            c.User.id == current_user.id,
                            c.User.workplace_restaurant_id.in_(allowed_workplaces),
                        ),
                    )
                    .distinct()
                )
                positions_query = positions_query.filter(c.Position.id.in_(accessible_position_ids))
            else:
                positions_query = positions_query.filter(False)
        positions = positions_query.all()

        subdivision_ids = sorted(
            {
                int(item.restaurant_subdivision_id)
                for item in positions
                if getattr(item, "restaurant_subdivision_id", None) is not None
            }
        )
        subdivisions_query = db.query(c.RestaurantSubdivision).order_by(c.RestaurantSubdivision.name.asc())
        if allowed_workplaces is not None:
            if subdivision_ids:
                subdivisions_query = subdivisions_query.filter(c.RestaurantSubdivision.id.in_(subdivision_ids))
            else:
                subdivisions_query = subdivisions_query.filter(False)
        subdivisions = subdivisions_query.all()

        return c.TimesheetOptionsResponse(
            restaurants=[
                {"id": item.id, "name": item.name or f"Restaurant #{item.id}"}
                for item in restaurants
                if item.id is not None
            ],
            subdivisions=[
                {"id": item.id, "name": item.name or f"Subdivision #{item.id}"}
                for item in subdivisions
                if item.id is not None
            ],
            positions=[
                {
                    "id": item.id,
                    "name": item.name or f"Position #{item.id}",
                    "restaurant_subdivision_id": getattr(item, "restaurant_subdivision_id", None),
                }
                for item in positions
                if item.id is not None
            ],
        ).model_dump(mode="json")

    payload = c.cached_reference_data(
        c.STAFF_TIMESHEET_OPTIONS_CACHE_SCOPE,
        cache_key,
        _load_timesheet_options,
        ttl_seconds=c.STAFF_TIMESHEET_OPTIONS_CACHE_TTL_SECONDS,
    )
    return c.TimesheetOptionsResponse.model_validate(payload)


@router.get("/timesheet/export")
def export_staff_timesheet(
    restaurant_id: int = Query(...),
    date_from: date = Query(...),
    date_to: date = Query(...),
    restaurant_subdivision_ids: Optional[List[int]] = Query(None),
    position_ids: Optional[List[int]] = Query(None),
    db: c.Session = Depends(c.get_db),
    current_user: c.User = Depends(c.get_current_user),
):
    c._ensure_timesheet_export(current_user)
    if date_to < date_from:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="date_to must be >= date_from")

    allowed_workplaces = c._get_allowed_workplace_ids(db, current_user)
    if allowed_workplaces is not None and restaurant_id not in allowed_workplaces:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    stream = c.build_timesheet_report(
        db,
        date_from=date_from,
        date_to=date_to,
        restaurant_id=restaurant_id,
        restaurant_subdivision_ids=restaurant_subdivision_ids,
        position_ids=position_ids,
    )
    label = date_from.isoformat() if date_from == date_to else f"{date_from.isoformat()}_{date_to.isoformat()}"
    filename = f"timesheet_{label}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export")
def export_staff_employees(
    payload: c.StaffEmployeesExportRequest,
    db: c.Session = Depends(c.get_db),
    current_user: c.User = Depends(c.get_current_user),
):
    c._ensure_staff_employees_export(current_user)

    raw_user_ids = payload.user_ids or []
    if not raw_user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_ids is required")
    if len(raw_user_ids) > 2000:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Too many employees to export")

    user_ids: list[int] = []
    seen_ids: set[int] = set()
    for value in raw_user_ids:
        try:
            uid = int(value)
        except Exception:
            continue
        if uid <= 0 or uid in seen_ids:
            continue
        seen_ids.add(uid)
        user_ids.append(uid)
    if not user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_ids is required")

    raw_column_ids = payload.column_ids or []
    column_ids: list[str] = []
    seen_columns: set[str] = set()
    for value in raw_column_ids:
        if value is None:
            continue
        cid = str(value).strip()
        if not cid or cid in seen_columns:
            continue
        if cid not in c.EMPLOYEE_EXPORT_COLUMNS:
            continue
        seen_columns.add(cid)
        column_ids.append(cid)
    if not column_ids:
        column_ids = list(c.EMPLOYEE_EXPORT_COLUMNS.keys())

    users_query = (
        db.query(c.User)
        .options(
            c.joinedload(c.User.company),
            c.joinedload(c.User.position),
            c.selectinload(c.User.restaurants),
        )
        .filter(c.User.id.in_(user_ids))
    )
    allowed_workplaces = c._get_allowed_workplace_ids(db, current_user)
    if allowed_workplaces is not None:
        if allowed_workplaces:
            users_query = users_query.filter(
                c.or_(
                    c.User.id == current_user.id,
                    c.User.workplace_restaurant_id.in_(allowed_workplaces),
                )
            )
        else:
            users_query = users_query.filter(c.User.id == current_user.id)

    users = users_query.all()
    users_by_id: dict[int, c.User] = {item.id: item for item in users if item and item.id is not None}
    inaccessible_user_ids = [uid for uid in user_ids if uid not in users_by_id]
    if inaccessible_user_ids:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied for one or more selected employees",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    ws.append([c.EMPLOYEE_EXPORT_COLUMNS[cid] for cid in column_ids])

    date_columns = {"hire_date", "fire_date", "birth_date"}

    for uid in user_ids:
        user = users_by_id.get(uid)
        if not user:
            continue
        ws.append([c._employee_export_value(user, cid) for cid in column_ids])
        row_idx = ws.max_row
        for col_idx, cid in enumerate(column_ids, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cid in date_columns and isinstance(cell.value, date):
                cell.number_format = "dd.mm.yyyy"
            if cid == "restaurants":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).coordinate}"

    c._style_sheet(ws, freeze="A2")
    c._autosize_sheet(ws, min_width=10.0, max_width=40.0)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"employees_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
