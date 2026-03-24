from __future__ import annotations

import logging
from decimal import Decimal
from io import BytesIO
import os
import re
from datetime import date, timedelta
from typing import List, Optional, Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy import and_, or_, func
from sqlalchemy.orm import Session, joinedload, selectinload, load_only, noload

from backend.bd.database import get_db
from backend.bd.models import (
    Attendance,
    User,
    Restaurant,
    RestaurantSubdivision,
    Role,
    Company,
    Position,
    EmployeeChangeEvent,
)
from backend.schemas import (
    AttendancePublic,
    StaffEmployeeDetailResponse,
    StaffEmployeeListResponse,
    StaffUserPublic,
    RestaurantRead,
    CompanyRead,
    RoleRead,
    PositionHierarchyNode,
    EmployeeCardPublic,
    EmployeeIikoSyncPreviewResponse,
    IikoSyncEmployeeSnapshot,
    EmployeeUpdateRequest,
    PhotoUploadResponse,
    TimesheetOptionsResponse,
)
from backend.services.iiko_staff import (
    add_user_to_iiko,
    fetch_iiko_employee_snapshot,
    fire_user_in_iiko,
    IikoIntegrationError,
)
from backend.services.payroll_export import _autosize_sheet, _style_sheet
from backend.services.reference_cache import cached_reference_data
from backend.services.timesheet_export import build_timesheet_report

logger = logging.getLogger(__name__)

try:  # pragma: no cover - dependency fallback for standalone usage
    from backend.utils import get_current_user, get_user_restaurant_ids, users_share_restaurant, hash_password
    from backend.services.employee_identity import build_employee_row_id
    from backend.services.permissions import (
        PermissionCode,
        ensure_permissions,
        has_permission,
        can_manage_user,
        ensure_can_manage_user,
        can_view_rate,
    )
    from backend.services.employee_changes import log_employee_changes, format_ref
except Exception as exc:  # pragma: no cover
    raise RuntimeError("Failed to import shared auth dependencies in staff employees router") from exc

from backend.services.s3 import generate_presigned_url, upload_employee_photo_with_url


router = APIRouter(prefix="/staff/employees", tags=["Staff employees"])
_RATE_FULL_ACCESS_LEVEL = 5
STAFF_REFERENCES_CACHE_SCOPE = "staff_employees:references"
STAFF_REFERENCES_CACHE_TTL_SECONDS = 45
STAFF_TIMESHEET_OPTIONS_CACHE_SCOPE = "staff_employees:timesheet_options"
STAFF_TIMESHEET_OPTIONS_CACHE_TTL_SECONDS = 45


def _month_bounds(day: date) -> tuple[date, date]:
    start = date(day.year, day.month, 1)
    if day.month == 12:
        end = date(day.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(day.year, day.month + 1, 1) - timedelta(days=1)
    return start, end


def _ids_scope_key(values: Optional[set[int]]) -> tuple:
    if values is None:
        return ("all",)
    if not values:
        return ("none",)
    return tuple(sorted(int(value) for value in values))


def _normalize_rate_value(value):
    if value is None:
        return None
    try:
        return str(Decimal(str(value)).quantize(Decimal("0.01")))
    except Exception:
        return str(value)


def _format_restaurant_refs(restaurants):
    if not restaurants:
        return []
    return [format_ref(item) for item in restaurants if item]


def _to_staff_public(
    user: User,
    viewer: Optional[User] = None,
    can_see_rate_override: Optional[bool] = None,
) -> StaffUserPublic:
    if can_see_rate_override is not None:
        can_see_rate = bool(can_see_rate_override)
    else:
        can_see_rate = can_view_rate(viewer, user) if viewer else True
    restaurants = [
       {
            "id": r.id,
            "name": (r.name or f"Restaurant #{r.id}"),
            "department_code": getattr(r, "department_code", None),
        }  # type: ignore[arg-type]
        for r in (user.restaurants or [])
        if r and r.id is not None
    ]
    restaurant_department_codes = [
        r.department_code for r in (user.restaurants or []) if getattr(r, "department_code", None)
    ]
    subdivision_id = getattr(getattr(user, "position", None), "restaurant_subdivision_id", None)
    subdivision_name = getattr(getattr(getattr(user, "position", None), "restaurant_subdivision", None), "name", None)
    subdivision_is_multi = bool(getattr(getattr(getattr(user, "position", None), "restaurant_subdivision", None), "is_multi", False))
    return StaffUserPublic(
        id=user.id,
        username=user.username,
        first_name=user.first_name,
        last_name=user.last_name,
        middle_name=user.middle_name,
        iiko_id=user.iiko_id,
        iiko_code=user.iiko_code,
        staff_code=user.staff_code,
        phone_number=user.phone_number,
        email=user.email,
        company_id=user.company.id if user.company else None,
        company_name=user.company.name if user.company else None,
        role_id=user.role_id,
        role_name=user.role.name if user.role else None,
        position_id=user.position_id,
        position_name=user.position.name if user.position else None,
        position_code=user.position.code if getattr(user, "position", None) else None,
        gender=user.gender,
        rate=float(user.rate) if can_see_rate and user.rate is not None else None,
        position_rate=float(user.position.rate) if can_see_rate and getattr(user, "position", None) and user.position.rate is not None else None,
        hire_date=user.hire_date,
        fire_date=user.fire_date,
        birth_date=user.birth_date,
        photo_key=user.photo_key,
        fired=bool(user.fired),
        is_formalized=bool(getattr(user, "is_formalized", False)),
        is_cis_employee=bool(getattr(user, "is_cis_employee", False)),
        confidential_data_consent=bool(getattr(user, "confidential_data_consent", False)),
        restaurants=restaurants,
        restaurant_ids=[item["id"] for item in restaurants],
        restaurant_department_codes=restaurant_department_codes,
        workplace_restaurant_id=getattr(user, "workplace_restaurant_id", None),
        individual_rate=float(user.individual_rate) if can_see_rate and user.individual_rate is not None else None,
        has_full_restaurant_access=bool(getattr(user, "has_full_restaurant_access", False)),
        rate_hidden=not can_see_rate,
        has_fingerprint=bool(getattr(user, "has_fingerprint", False)),
        restaurant_subdivision_id=subdivision_id,
        restaurant_subdivision_name=subdivision_name,
        restaurant_subdivision_is_multi=subdivision_is_multi,
    )


def _to_employee_card(user: User, viewer: Optional[User] = None) -> EmployeeCardPublic:
    can_see_rate = can_view_rate(viewer, user) if viewer else True
    photo_url = None
    if user.photo_key:
        try:
            photo_url = generate_presigned_url(user.photo_key)
        except Exception:
            logger.warning("Failed to build photo URL for user %s", user.id, exc_info=True)
            photo_url = None

    return EmployeeCardPublic(
        id=user.id,
        username=user.username,
        first_name=user.first_name,
        last_name=user.last_name,
        middle_name=user.middle_name,
        iiko_code=user.iiko_code,
        iiko_id=user.iiko_id,
        company_id=user.company.id if user.company else None,
        company_name=user.company.name if user.company else None,
        role_id=user.role_id,
        role_name=user.role.name if user.role else None,
        position_id=user.position_id,
        position_name=user.position.name if user.position else None,
        rate=float(user.rate) if can_see_rate and user.rate is not None else None,
        hire_date=user.hire_date,
        fire_date=user.fire_date,
        fired=bool(user.fired),
        staff_code=user.staff_code,
        gender=user.gender,
        birth_date=user.birth_date,
        phone_number=user.phone_number,
        email=user.email,
        photo_key=user.photo_key,
        photo_url=photo_url,
        is_cis_employee=bool(getattr(user, "is_cis_employee", False)),
        confidential_data_consent=bool(getattr(user, "confidential_data_consent", False)),
        is_formalized=bool(getattr(user, "is_formalized", False)),
        workplace_restaurant_id=getattr(user, "workplace_restaurant_id", None),
        individual_rate=float(user.individual_rate) if can_see_rate and user.individual_rate is not None else None,
        iiko_sync_error=getattr(user, "iiko_sync_error", None),
        rate_hidden=not can_see_rate,
        has_fingerprint=bool(getattr(user, "has_fingerprint", False)),
    )


def _get_duplicate_photo_url(user: User) -> Optional[str]:
    if not user.photo_key:
        return None
    try:
        return generate_presigned_url(user.photo_key)
    except Exception:
        logger.warning("Failed to build photo URL for duplicate employee %s", user.id, exc_info=True)
        return None


def _get_duplicate_fired_comment(db: Session, user: User) -> Optional[str]:
    if not user.fired:
        return None
    event = (
        db.query(EmployeeChangeEvent)
        .filter(
            EmployeeChangeEvent.user_id == user.id,
            EmployeeChangeEvent.field == "fired",
            EmployeeChangeEvent.new_value.in_(["True", "true", "1"]),
        )
        .order_by(EmployeeChangeEvent.created_at.desc(), EmployeeChangeEvent.id.desc())
        .first()
    )
    return event.comment if event else None


def _can_view_duplicate_employee(db: Session, viewer: User, target: User) -> bool:
    if viewer.id == target.id:
        return True
    if not users_share_restaurant(db, viewer, target.id):
        return False
    return any(
        has_permission(viewer, code)
        for code in (
            PermissionCode.STAFF_VIEW_SENSITIVE,
            PermissionCode.STAFF_MANAGE_SUBORDINATES,
            PermissionCode.STAFF_MANAGE_ALL,
            PermissionCode.STAFF_EMPLOYEES_VIEW,
            PermissionCode.EMPLOYEES_CARD_VIEW,
        )
    )


def _build_duplicate_employee_payload(db: Session, user: User, *, include_id: bool) -> dict:
    position_name = user.position.name if user.position else None
    workplace_name = user.workplace_restaurant.name if user.workplace_restaurant else None
    payload = {
        "first_name": user.first_name or "",
        "last_name": user.last_name or "",
        "middle_name": user.middle_name or "",
        "birth_date": user.birth_date.isoformat() if user.birth_date else None,
        "hire_date": user.hire_date.isoformat() if user.hire_date else None,
        "fire_date": user.fire_date.isoformat() if user.fire_date else None,
        "position": position_name,
        "workplace": workplace_name,
        "fired": bool(user.fired),
        "fired_comment": _get_duplicate_fired_comment(db, user),
        "photo_url": _get_duplicate_photo_url(user),
    }
    if include_id:
        payload["id"] = user.id
    return payload


def _attendance_to_public(
    attendance: Attendance,
    viewer: Optional[User] = None,
    target: Optional[User] = None,
) -> AttendancePublic:
    can_see_rate = True
    if viewer is not None and target is not None:
        can_see_rate = can_view_rate(viewer, target)
    return AttendancePublic(
        id=attendance.id,
        user_id=attendance.user_id,
        position_id=attendance.position_id,
        position_name=attendance.position.name if attendance.position else None,
        restaurant_id=attendance.restaurant_id,
        restaurant_name=attendance.restaurant.name if attendance.restaurant else None,
        rate=float(attendance.rate) if can_see_rate and attendance.rate is not None else None,
        pay_amount=float(attendance.pay_amount) if can_see_rate and attendance.pay_amount is not None else None,
        open_date=attendance.open_date,
        open_time=attendance.open_time,
        close_date=attendance.close_date,
        close_time=attendance.close_time,
        duration_minutes=attendance.duration_minutes,
        night_minutes=attendance.night_minutes or 0,
    )


def _attendance_range_condition(date_from: date, date_to: date):
    # фильтруем по дате открытия смены
    return and_(
        Attendance.open_date >= date_from,
        Attendance.open_date <= date_to,
    )


def _ensure_staff_view(user: User) -> None:
    ensure_permissions(
        user,
        PermissionCode.STAFF_VIEW_SENSITIVE,
        PermissionCode.STAFF_MANAGE_SUBORDINATES,
        PermissionCode.STAFF_MANAGE_ALL,
        PermissionCode.STAFF_EMPLOYEES_VIEW,
        PermissionCode.EMPLOYEES_CARD_VIEW,
    )


def _role_level_for_rate(user: Optional[User]) -> Optional[int]:
    if not user:
        return None
    role = getattr(user, "role", None)
    if role is None:
        role = getattr(getattr(user, "position", None), "role", None)
    if role is None or getattr(role, "level", None) is None:
        return None
    try:
        return int(role.level)
    except Exception:
        return None


def _ensure_timesheet_export(user: User) -> None:
    _ensure_staff_view(user)
    ensure_permissions(user, PermissionCode.TIMESHEET_EXPORT)


def _get_allowed_workplace_ids(db: Session, viewer: User) -> Optional[set[int]]:
    allowed_restaurants = get_user_restaurant_ids(db, viewer)
    if allowed_restaurants is None:
        return None
    allowed_workplaces = set(allowed_restaurants)
    if getattr(viewer, "workplace_restaurant_id", None):
        allowed_workplaces.add(viewer.workplace_restaurant_id)
    return allowed_workplaces


def _validate_iiko_restaurant_scope(
    db: Session,
    viewer: User,
    *,
    sync_restaurant_id: Optional[int],
    department_restaurant_ids: Optional[list[int]],
) -> None:
    allowed_restaurants = get_user_restaurant_ids(db, viewer)
    if allowed_restaurants is None:
        return

    requested_ids: set[int] = set()
    if sync_restaurant_id:
        requested_ids.add(int(sync_restaurant_id))
    for raw in department_restaurant_ids or []:
        try:
            rid = int(raw)
        except Exception:
            continue
        if rid > 0:
            requested_ids.add(rid)

    invalid_ids = requested_ids.difference(allowed_restaurants)
    if invalid_ids:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access to the selected restaurants is not allowed")


def _normalize_role_key(role: Optional[Role]) -> str:
    name = getattr(role, "name", None) or ""
    return name.strip().lower().replace(" ", "").replace("-", "")


def _can_edit_iiko_id(viewer: User) -> bool:
    if has_permission(viewer, PermissionCode.SYSTEM_ADMIN):
        return True
    role_key = _normalize_role_key(getattr(viewer, "role", None))
    return role_key in {
        "суперадмин",
        "системныйадмин",
        "системныйадминистратор",
        "superadmin",
        "systemadmin",
        "systemadministrator",
    }


def _ensure_staff_employees_export(user: User) -> None:
    _ensure_staff_view(user)
    ensure_permissions(user, PermissionCode.STAFF_EMPLOYEES_EXPORT)


EMPLOYEE_EXPORT_COLUMNS: dict[str, str] = {
    "staff_code": "Табельный номер",
    "full_name": "ФИО",
    "phone_number": "Телефон",
    "company_name": "Компания",
    "position_name": "Должность",
    "iiko_id": "Код сотрудника (Айко)",
    "gender": "Пол",
    "hire_date": "Дата найма",
    "fire_date": "Дата увольнения",
    "birth_date": "Дата рождения",
    "is_cis_employee": "Сотрудник СНГ",
    "restaurants": "Рестораны",
    "status": "Статус",
}


def _format_employee_phone(value: Optional[str]) -> str:
    if value is None or value == "":
        return ""
    digits = re.sub(r"\\D", "", str(value))
    if not digits:
        return str(value)
    normalized = digits
    if normalized.startswith("8"):
        normalized = f"7{normalized[1:]}"
    if not normalized.startswith("7"):
        normalized = f"7{normalized}"
    normalized = normalized[:11]
    if len(normalized) < 11:
        return str(value)
    area = normalized[1:4]
    first = normalized[4:7]
    second = normalized[7:9]
    third = normalized[9:11]
    return f"+7({area})-{first}-{second}-{third}"


def _format_employee_gender(value: Optional[str]) -> str:
    if value == "male":
        return "Мужской"
    if value == "female":
        return "Женский"
    return ""


def _employee_export_full_name(user: User) -> str:
    last_name = getattr(user, "last_name", None) or ""
    first_name = getattr(user, "first_name", None) or ""
    parts = [part for part in [last_name, first_name] if part]
    if parts:
        return " ".join(parts)
    middle_name = getattr(user, "middle_name", None) or ""
    parts = [part for part in [last_name, first_name, middle_name] if part]
    if parts:
        return " ".join(parts)
    return getattr(user, "username", None) or ""


def _employee_export_restaurants(user: User) -> str:
    restaurants = getattr(user, "restaurants", None) or []
    names: list[str] = []
    for restaurant in restaurants:
        if not restaurant:
            continue
        name = getattr(restaurant, "name", None)
        if name:
            names.append(str(name))
            continue
        rid = getattr(restaurant, "id", None)
        if rid is not None:
            names.append(f"ID {rid}")
    return ", ".join([n for n in names if n])


def _employee_export_value(user: User, column_id: str) -> Any:
    if column_id == "staff_code":
        return getattr(user, "staff_code", None) or ""
    if column_id == "full_name":
        return _employee_export_full_name(user)
    if column_id == "phone_number":
        return _format_employee_phone(getattr(user, "phone_number", None))
    if column_id == "company_name":
        company = getattr(user, "company", None)
        return getattr(company, "name", None) or ""
    if column_id == "position_name":
        position = getattr(user, "position", None)
        return getattr(position, "name", None) or ""
    if column_id == "iiko_id":
        return getattr(user, "iiko_id", None) or ""
    if column_id == "gender":
        return _format_employee_gender(getattr(user, "gender", None))
    if column_id == "hire_date":
        return getattr(user, "hire_date", None) or None
    if column_id == "fire_date":
        return getattr(user, "fire_date", None) or None
    if column_id == "birth_date":
        return getattr(user, "birth_date", None) or None
    if column_id == "is_cis_employee":
        return "Да" if bool(getattr(user, "is_cis_employee", False)) else "Нет"
    if column_id == "restaurants":
        return _employee_export_restaurants(user)
    if column_id == "status":
        return "Уволен" if bool(getattr(user, "fired", False)) else "Активен"
    return ""


class StaffEmployeesExportRequest(BaseModel):
    user_ids: List[int] = Field(default_factory=list, description="Ordered user IDs in current UI scope")
    column_ids: List[str] = Field(default_factory=list, description="Ordered column IDs to include")


class StaffEmployeesReferencePayload(BaseModel):
    restaurants: List[RestaurantRead] = Field(default_factory=list)
    companies: List[CompanyRead] = Field(default_factory=list)
    roles: List[RoleRead] = Field(default_factory=list)
    positions: List[PositionHierarchyNode] = Field(default_factory=list)


class StaffEmployeesBootstrapResponse(BaseModel):
    items: List[StaffUserPublic] = Field(default_factory=list)
    references: StaffEmployeesReferencePayload = Field(default_factory=StaffEmployeesReferencePayload)


def _can_edit_user(db: Session, viewer: User, target: User) -> bool:
    if has_permission(viewer, PermissionCode.STAFF_MANAGE_ALL):
        return True
    if has_permission(viewer, PermissionCode.STAFF_MANAGE_SUBORDINATES):
        return can_manage_user(viewer, target)
    if (
        has_permission(viewer, PermissionCode.STAFF_EMPLOYEES_MANAGE)
        or has_permission(viewer, PermissionCode.EMPLOYEES_CARD_MANAGE)
    ) and users_share_restaurant(db, viewer, target.id):
        return True
    return False


def _ensure_can_edit_user(db: Session, viewer: User, target: User) -> None:
    if _can_edit_user(db, viewer, target):
        return
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Access denied: insufficient permissions to manage this employee",
    )


def _can_sync_user_iiko(db: Session, viewer: User, target: User) -> bool:
    if _can_edit_user(db, viewer, target):
        return True

    has_sync_permissions = (
        has_permission(viewer, PermissionCode.STAFF_EMPLOYEES_IIKO_SYNC)
        or has_permission(viewer, PermissionCode.IIKO_MANAGE)
    )
    if not has_sync_permissions:
        return False

    return users_share_restaurant(db, viewer, target.id)


def _ensure_can_sync_user_iiko(db: Session, viewer: User, target: User) -> None:
    if _can_sync_user_iiko(db, viewer, target):
        return
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Access denied: insufficient permissions to sync this employee with iiko",
    )


def _can_load_restaurant_references(user: User) -> bool:
    return any(
        has_permission(user, code)
        for code in (
            PermissionCode.SYSTEM_ADMIN,
            PermissionCode.RESTAURANTS_VIEW,
            PermissionCode.RESTAURANTS_MANAGE,
            PermissionCode.RESTAURANTS_SETTINGS_VIEW,
            PermissionCode.RESTAURANTS_SETTINGS_MANAGE,
            PermissionCode.STAFF_PORTAL_ACCESS,
        )
    )


def _can_load_company_references(user: User) -> bool:
    return any(
        has_permission(user, code)
        for code in (
            PermissionCode.SYSTEM_ADMIN,
            PermissionCode.COMPANIES_VIEW,
            PermissionCode.COMPANIES_MANAGE,
        )
    )


def _can_load_role_references(user: User) -> bool:
    return any(
        has_permission(user, code)
        for code in (
            PermissionCode.SYSTEM_ADMIN,
            PermissionCode.ROLES_MANAGE,
            PermissionCode.STAFF_ROLES_ASSIGN,
        )
    )


def _can_load_position_references(user: User) -> bool:
    return any(
        has_permission(user, code)
        for code in (
            PermissionCode.SYSTEM_ADMIN,
            PermissionCode.POSITIONS_MANAGE,
            PermissionCode.POSITIONS_EDIT,
            PermissionCode.POSITIONS_RATE_MANAGE,
        )
    )


def _position_hierarchy_payload(position: Position) -> PositionHierarchyNode:
    return PositionHierarchyNode(
        id=position.id,
        name=position.name,
        code=position.code,
        role_id=position.role_id,
        role_name=position.role.name if position.role else None,
        parent_id=position.parent_id,
        hierarchy_level=position.hierarchy_level,
        rate=position.rate,
        payment_format_id=position.payment_format_id,
        payment_format_name=position.payment_format.name if position.payment_format else None,
        hours_per_shift=position.hours_per_shift,
        monthly_shift_norm=position.monthly_shift_norm,
        restaurant_subdivision_id=position.restaurant_subdivision_id,
        restaurant_subdivision_name=position.restaurant_subdivision.name if position.restaurant_subdivision else None,
        night_bonus_enabled=bool(getattr(position, "night_bonus_enabled", False)),
        night_bonus_percent=getattr(position, "night_bonus_percent", None),
    )


def _load_staff_references(db: Session, current_user: User) -> StaffEmployeesReferencePayload:
    can_load_restaurants = _can_load_restaurant_references(current_user)
    can_load_companies = _can_load_company_references(current_user)
    can_load_roles = _can_load_role_references(current_user)
    can_load_positions = _can_load_position_references(current_user)
    allowed_restaurants = get_user_restaurant_ids(db, current_user) if can_load_restaurants else None

    cache_key = (
        int(current_user.id),
        bool(can_load_restaurants),
        bool(can_load_companies),
        bool(can_load_roles),
        bool(can_load_positions),
        _ids_scope_key(allowed_restaurants) if can_load_restaurants else ("skip",),
    )

    def _load_references() -> dict:
        restaurants_payload: list[RestaurantRead] = []
        companies_payload: list[CompanyRead] = []
        roles_payload: list[RoleRead] = []
        positions_payload: list[PositionHierarchyNode] = []

        if can_load_restaurants:
            restaurants_query = db.query(Restaurant)
            if allowed_restaurants is not None:
                if allowed_restaurants:
                    restaurants_query = restaurants_query.filter(Restaurant.id.in_(allowed_restaurants))
                else:
                    restaurants_query = restaurants_query.filter(False)
            restaurants = restaurants_query.order_by(Restaurant.id.asc()).all()
            restaurants_payload = [RestaurantRead.model_validate(item) for item in restaurants]

        if can_load_companies:
            companies = db.query(Company).order_by(Company.id.asc()).all()
            companies_payload = [CompanyRead.model_validate(item) for item in companies]

        if can_load_roles:
            roles = db.query(Role).order_by(func.lower(Role.name).nullslast(), Role.id.asc()).all()
            roles_payload = [RoleRead.model_validate(item) for item in roles]

        if can_load_positions:
            positions = (
                db.query(Position)
                .options(
                    joinedload(Position.role).load_only(Role.id, Role.name),
                    joinedload(Position.payment_format),
                    joinedload(Position.restaurant_subdivision).load_only(
                        RestaurantSubdivision.id,
                        RestaurantSubdivision.name,
                    ),
                )
                .order_by(Position.hierarchy_level.asc(), func.lower(Position.name).nullslast(), Position.id.asc())
                .all()
            )
            positions_payload = [_position_hierarchy_payload(item) for item in positions]

        return StaffEmployeesReferencePayload(
            restaurants=restaurants_payload,
            companies=companies_payload,
            roles=roles_payload,
            positions=positions_payload,
        ).model_dump(mode="json")

    payload = cached_reference_data(
        STAFF_REFERENCES_CACHE_SCOPE,
        cache_key,
        _load_references,
        ttl_seconds=STAFF_REFERENCES_CACHE_TTL_SECONDS,
    )
    return StaffEmployeesReferencePayload.model_validate(payload)


def _list_staff_employee_items(
    db: Session,
    current_user: User,
    *,
    q: Optional[str] = None,
    include_fired: bool = False,
    only_fired: bool = False,
    restaurant_id: Optional[int] = None,
    hire_date_from: Optional[date] = None,
    hire_date_to: Optional[date] = None,
    fire_date_from: Optional[date] = None,
    fire_date_to: Optional[date] = None,
    limit: int = 1000,
) -> List[StaffUserPublic]:
    _ensure_staff_view(current_user)
    viewer_can_view_all_rates = has_permission(current_user, PermissionCode.STAFF_RATE_VIEW_ALL)
    viewer_rate_level = _role_level_for_rate(current_user)

    def _can_see_rate_for_target(target: User) -> bool:
        if viewer_can_view_all_rates:
            return True
        if current_user.id == target.id:
            return True
        if viewer_rate_level is None:
            return False
        if viewer_rate_level >= _RATE_FULL_ACCESS_LEVEL:
            return True
        target_level = _role_level_for_rate(target)
        if target_level is None:
            target_level = -1
        return viewer_rate_level > target_level

    query = db.query(User).options(
        noload(User.permission_links),
        noload(User.direct_permissions),
        noload(User.medical_check_records),
        noload(User.cis_document_records),
        noload(User.training_events),
        load_only(
            User.id,
            User.username,
            User.first_name,
            User.last_name,
            User.middle_name,
            User.iiko_id,
            User.iiko_code,
            User.staff_code,
            User.phone_number,
            User.email,
            User.company_id,
            User.role_id,
            User.position_id,
            User.gender,
            User.rate,
            User.hire_date,
            User.fire_date,
            User.birth_date,
            User.photo_key,
            User.fired,
            User.is_cis_employee,
            User.is_formalized,
            User.confidential_data_consent,
            User.workplace_restaurant_id,
            User.individual_rate,
            User.has_full_restaurant_access,
            User.has_fingerprint,
        ),
        joinedload(User.role).load_only(Role.id, Role.name, Role.level),
        joinedload(User.role).noload(Role.permission_links),
        joinedload(User.role).noload(Role.permissions),
        joinedload(User.company).load_only(Company.id, Company.name),
        joinedload(User.position)
        .load_only(
            Position.id,
            Position.name,
            Position.code,
            Position.rate,
            Position.restaurant_subdivision_id,
            Position.role_id,
        ),
        joinedload(User.position).noload(Position.permission_links),
        joinedload(User.position).noload(Position.permissions),
        joinedload(User.position).noload(Position.training_requirements),
        joinedload(User.position)
        .joinedload(Position.role)
        .load_only(Role.id, Role.name, Role.level),
        joinedload(User.position).joinedload(Position.role).noload(Role.permission_links),
        joinedload(User.position).joinedload(Position.role).noload(Role.permissions),
        joinedload(User.position)
        .joinedload(Position.restaurant_subdivision)
        .load_only(
            RestaurantSubdivision.id,
            RestaurantSubdivision.name,
            RestaurantSubdivision.is_multi,
        ),
        selectinload(User.restaurants).load_only(Restaurant.id, Restaurant.name, Restaurant.department_code),
    )
    if only_fired:
        query = query.filter(User.fired == True)
    elif not include_fired:
        query = query.filter(User.fired == False)
    if q:
        like = f"%{q.lower()}%"
        query = query.filter(
            or_(
                func.lower(User.username).like(like),
                func.lower(func.coalesce(User.first_name, "")).like(like),
                func.lower(func.coalesce(User.last_name, "")).like(like),
                func.lower(func.coalesce(User.middle_name, "")).like(like),
                func.lower(func.coalesce(User.staff_code, "")).like(like),
                func.lower(func.coalesce(User.phone_number, "")).like(like),
            )
        )
    if hire_date_from:
        query = query.filter(User.hire_date >= hire_date_from)
    if hire_date_to:
        query = query.filter(User.hire_date <= hire_date_to)
    if fire_date_from:
        query = query.filter(User.fire_date >= fire_date_from)
    if fire_date_to:
        query = query.filter(User.fire_date <= fire_date_to)
    if restaurant_id is not None:
        query = query.filter(User.workplace_restaurant_id == restaurant_id)

    allowed_workplaces = _get_allowed_workplace_ids(db, current_user)
    if allowed_workplaces is not None:
        if restaurant_id is not None and restaurant_id not in allowed_workplaces:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
        if allowed_workplaces:
            query = query.filter(
                or_(
                    User.id == current_user.id,
                    User.workplace_restaurant_id.in_(allowed_workplaces),
                )
            )
        else:
            query = query.filter(User.id == current_user.id)
    users = (
        query.order_by(
            func.lower(User.last_name).nullslast(),
            func.lower(User.first_name).nullslast(),
            User.id.asc(),
        )
        .limit(limit)
        .all()
    )

    return [
        _to_staff_public(u, current_user, can_see_rate_override=_can_see_rate_for_target(u))
        for u in users
    ]


@router.get("", response_model=StaffEmployeeListResponse)
def list_staff_employees(
    q: Optional[str] = Query(None, description="Search by username/first_name/last_name/staff_code"),
    include_fired: bool = Query(False),
    only_fired: bool = Query(False),
    restaurant_id: Optional[int] = Query(None, ge=1),
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    fire_date_from: Optional[date] = Query(None),
    fire_date_to: Optional[date] = Query(None),
    limit: int = Query(1000, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StaffEmployeeListResponse:
    items = _list_staff_employee_items(
        db,
        current_user,
        q=q,
        include_fired=include_fired,
        only_fired=only_fired,
        restaurant_id=restaurant_id,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        fire_date_from=fire_date_from,
        fire_date_to=fire_date_to,
        limit=limit,
    )
    return StaffEmployeeListResponse(items=items)


@router.get("/bootstrap", response_model=StaffEmployeesBootstrapResponse)
def staff_employees_bootstrap(
    q: Optional[str] = Query(None, description="Search by username/first_name/last_name/staff_code"),
    include_fired: bool = Query(False),
    only_fired: bool = Query(False),
    restaurant_id: Optional[int] = Query(None, ge=1),
    hire_date_from: Optional[date] = Query(None),
    hire_date_to: Optional[date] = Query(None),
    fire_date_from: Optional[date] = Query(None),
    fire_date_to: Optional[date] = Query(None),
    limit: int = Query(1000, ge=1, le=1000),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StaffEmployeesBootstrapResponse:
    items = _list_staff_employee_items(
        db,
        current_user,
        q=q,
        include_fired=include_fired,
        only_fired=only_fired,
        restaurant_id=restaurant_id,
        hire_date_from=hire_date_from,
        hire_date_to=hire_date_to,
        fire_date_from=fire_date_from,
        fire_date_to=fire_date_to,
        limit=limit,
    )
    references = _load_staff_references(db, current_user)
    return StaffEmployeesBootstrapResponse(items=items, references=references)


@router.get("/timesheet/options", response_model=TimesheetOptionsResponse)
def list_timesheet_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_timesheet_export(current_user)
    allowed_workplaces = _get_allowed_workplace_ids(db, current_user)
    cache_key = _ids_scope_key(allowed_workplaces)

    def _load_timesheet_options() -> dict:
        restaurants_query = db.query(Restaurant).order_by(func.lower(Restaurant.name).nullslast(), Restaurant.id.asc())
        if allowed_workplaces is not None:
            if allowed_workplaces:
                restaurants_query = restaurants_query.filter(Restaurant.id.in_(allowed_workplaces))
            else:
                restaurants_query = restaurants_query.filter(False)

        restaurants = restaurants_query.all()
        subdivisions = db.query(RestaurantSubdivision).order_by(RestaurantSubdivision.name.asc()).all()
        positions = (
            db.query(Position)
            .options(joinedload(Position.restaurant_subdivision))
            .order_by(Position.name.asc())
            .all()
        )

        return TimesheetOptionsResponse(
            restaurants=[
                {"id": item.id, "name": item.name or f"Restaurant #{item.id}"}
                for item in restaurants
                if item.id is not None
            ],
            subdivisions=[
                {"id": item.id, "name": item.name or f"Subdivision #{item.id}"}
                for item in subdivisions
                if item.id is not None
            ],
            positions=[
                {
                    "id": item.id,
                    "name": item.name or f"Position #{item.id}",
                    "restaurant_subdivision_id": getattr(item, "restaurant_subdivision_id", None),
                }
                for item in positions
                if item.id is not None
            ],
        ).model_dump(mode="json")

    payload = cached_reference_data(
        STAFF_TIMESHEET_OPTIONS_CACHE_SCOPE,
        cache_key,
        _load_timesheet_options,
        ttl_seconds=STAFF_TIMESHEET_OPTIONS_CACHE_TTL_SECONDS,
    )
    return TimesheetOptionsResponse.model_validate(payload)


@router.get("/timesheet/export")
def export_staff_timesheet(
    restaurant_id: int = Query(...),
    date_from: date = Query(...),
    date_to: date = Query(...),
    restaurant_subdivision_ids: Optional[List[int]] = Query(None),
    position_ids: Optional[List[int]] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_timesheet_export(current_user)
    if date_to < date_from:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="date_to must be >= date_from")

    allowed_workplaces = _get_allowed_workplace_ids(db, current_user)
    if allowed_workplaces is not None and restaurant_id not in allowed_workplaces:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    stream = build_timesheet_report(
        db,
        date_from=date_from,
        date_to=date_to,
        restaurant_id=restaurant_id,
        restaurant_subdivision_ids=restaurant_subdivision_ids,
        position_ids=position_ids,
    )
    label = date_from.isoformat() if date_from == date_to else f"{date_from.isoformat()}_{date_to.isoformat()}"
    filename = f"timesheet_{label}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export")
def export_staff_employees(
    payload: StaffEmployeesExportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _ensure_staff_employees_export(current_user)

    raw_user_ids = payload.user_ids or []
    if not raw_user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_ids is required")
    if len(raw_user_ids) > 2000:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Too many employees to export")

    user_ids: list[int] = []
    seen_ids: set[int] = set()
    for value in raw_user_ids:
        try:
            uid = int(value)
        except Exception:
            continue
        if uid <= 0 or uid in seen_ids:
            continue
        seen_ids.add(uid)
        user_ids.append(uid)
    if not user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_ids is required")

    raw_column_ids = payload.column_ids or []
    column_ids: list[str] = []
    seen_columns: set[str] = set()
    for value in raw_column_ids:
        if value is None:
            continue
        cid = str(value).strip()
        if not cid or cid in seen_columns:
            continue
        if cid not in EMPLOYEE_EXPORT_COLUMNS:
            continue
        seen_columns.add(cid)
        column_ids.append(cid)
    if not column_ids:
        column_ids = list(EMPLOYEE_EXPORT_COLUMNS.keys())

    users = (
        db.query(User)
        .options(
            joinedload(User.company),
            joinedload(User.position),
            selectinload(User.restaurants),
        )
        .filter(User.id.in_(user_ids))
        .all()
    )
    users_by_id: dict[int, User] = {item.id: item for item in users if item and item.id is not None}

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"
    ws.append([EMPLOYEE_EXPORT_COLUMNS[cid] for cid in column_ids])

    date_columns = {"hire_date", "fire_date", "birth_date"}

    for uid in user_ids:
        user = users_by_id.get(uid)
        if not user:
            continue
        ws.append([_employee_export_value(user, cid) for cid in column_ids])
        row_idx = ws.max_row
        for col_idx, cid in enumerate(column_ids, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cid in date_columns and isinstance(cell.value, date):
                cell.number_format = "dd.mm.yyyy"
            if cid == "restaurants":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).coordinate}"

    _style_sheet(ws, freeze="A2")
    _autosize_sheet(ws, min_width=10.0, max_width=40.0)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"employees_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/{user_id}", response_model=StaffEmployeeDetailResponse)
def get_staff_employee_detail(
    user_id: int,
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> StaffEmployeeDetailResponse:
    target = (
        db.query(User)
        .options(
            joinedload(User.role),
            joinedload(User.company),
            joinedload(User.position).joinedload(Position.payment_format),
            joinedload(User.position).joinedload(Position.restaurant_subdivision),
            joinedload(User.restaurants),
        )
        .filter(User.id == user_id)
        .first()
    )
    if not target:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    if current_user.id != user_id:
        _ensure_staff_view(current_user)
        if not users_share_restaurant(db, current_user, user_id):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")

    if not date_from or not date_to:
        start, end = _month_bounds(date.today())
        date_from = date_from or start
        date_to = date_to or end
    if date_from > date_to:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="date_from must be before or equal to date_to")

    condition = _attendance_range_condition(date_from, date_to)
    attendance_query = (
        db.query(Attendance)
        .options(
            joinedload(Attendance.position),
            joinedload(Attendance.restaurant),
        )
        .filter(Attendance.user_id == user_id)
        .filter(condition)
    )

    if current_user.id != user_id:
        allowed_restaurants = get_user_restaurant_ids(db, current_user)
        if allowed_restaurants is not None:
            if allowed_restaurants:
                attendance_query = attendance_query.filter(Attendance.restaurant_id.in_(allowed_restaurants))
            else:
                attendance_query = attendance_query.filter(False)

    rows = (
        attendance_query
        .order_by(Attendance.open_date.asc(), Attendance.open_time.asc())
        .all()
    )

    attendances = [_attendance_to_public(a, current_user, target) for a in rows]
    return StaffEmployeeDetailResponse(
        user=_to_staff_public(target, current_user),
        attendances=attendances,
        date_from=date_from,
        date_to=date_to,
    )


@router.get("/{user_id}/iiko-sync-preview", response_model=EmployeeIikoSyncPreviewResponse)
def get_employee_iiko_sync_preview(
    user_id: int,
    sync_restaurant_id: Optional[int] = Query(None, ge=1),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EmployeeIikoSyncPreviewResponse:
    db_user = (
        db.query(User)
        .options(
            joinedload(User.workplace_restaurant),
            joinedload(User.position),
            selectinload(User.restaurants),
        )
        .filter(User.id == user_id)
        .first()
    )
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    _ensure_can_sync_user_iiko(db, current_user, db_user)
    ensure_permissions(
        current_user,
        PermissionCode.STAFF_EMPLOYEES_IIKO_SYNC,
        PermissionCode.IIKO_MANAGE,
    )
    _validate_iiko_restaurant_scope(
        db,
        current_user,
        sync_restaurant_id=sync_restaurant_id,
        department_restaurant_ids=None,
    )

    local_restaurants = [row for row in (db_user.restaurants or []) if row and row.id is not None]
    if getattr(db_user, "workplace_restaurant", None) is not None and getattr(db_user.workplace_restaurant, "id", None):
        workplace_id = int(db_user.workplace_restaurant.id)
        if not any(int(row.id) == workplace_id for row in local_restaurants):
            local_restaurants.append(db_user.workplace_restaurant)
    local_department_codes: list[str] = []
    for row in local_restaurants:
        code = (getattr(row, "department_code", None) or "").strip()
        if code and code not in local_department_codes:
            local_department_codes.append(code)

    local_snapshot = IikoSyncEmployeeSnapshot(
        first_name=db_user.first_name,
        middle_name=db_user.middle_name,
        last_name=db_user.last_name,
        position_name=getattr(getattr(db_user, "position", None), "name", None),
        position_code=getattr(getattr(db_user, "position", None), "code", None),
        staff_code=db_user.staff_code,
        iiko_code=db_user.iiko_code,
        iiko_id=db_user.iiko_id,
        workplace_restaurant_id=getattr(db_user, "workplace_restaurant_id", None),
        workplace_restaurant_name=getattr(getattr(db_user, "workplace_restaurant", None), "name", None),
        department_code=getattr(getattr(db_user, "workplace_restaurant", None), "department_code", None),
        restaurant_ids=[int(row.id) for row in local_restaurants],
        restaurant_names=[
            row.name or f"Restaurant #{row.id}"
            for row in local_restaurants
        ],
        department_codes=local_department_codes,
    )

    iiko_snapshot = None
    iiko_error = None
    try:
        raw_iiko_snapshot = fetch_iiko_employee_snapshot(
            db,
            db_user,
            sync_restaurant_id=sync_restaurant_id,
        )
        if raw_iiko_snapshot:
            iiko_position_code = (raw_iiko_snapshot.get("main_role_code") or "").strip() or None
            iiko_position_name = None
            if iiko_position_code:
                iiko_position = (
                    db.query(Position)
                    .filter(func.lower(Position.code) == iiko_position_code.lower())
                    .first()
                )
                if iiko_position:
                    iiko_position_name = iiko_position.name

            department_code = (raw_iiko_snapshot.get("department_code") or "").strip() or None
            raw_department_codes = raw_iiko_snapshot.get("department_codes") or []
            if not isinstance(raw_department_codes, list):
                raw_department_codes = [raw_department_codes]

            department_codes: list[str] = []
            for raw_code in raw_department_codes:
                text = str(raw_code or "").strip()
                if text and text not in department_codes:
                    department_codes.append(text)
            if department_code and department_code not in department_codes:
                department_codes.insert(0, department_code)

            iiko_restaurants: list[Restaurant] = []
            for code in department_codes:
                restaurant = (
                    db.query(Restaurant)
                    .filter(Restaurant.department_code == code)
                    .first()
                )
                if not restaurant:
                    restaurant = (
                        db.query(Restaurant)
                        .filter(func.lower(Restaurant.department_code) == code.lower())
                        .first()
                    )
                if not restaurant:
                    continue
                if any(existing.id == restaurant.id for existing in iiko_restaurants):
                    continue
                iiko_restaurants.append(restaurant)

            primary_restaurant = None
            if department_code:
                primary_restaurant = next(
                    (row for row in iiko_restaurants if (getattr(row, "department_code", None) or "").strip() == department_code),
                    None,
                )
            if primary_restaurant is None and iiko_restaurants:
                primary_restaurant = iiko_restaurants[0]

            iiko_snapshot = IikoSyncEmployeeSnapshot(
                first_name=raw_iiko_snapshot.get("first_name") or None,
                middle_name=raw_iiko_snapshot.get("middle_name") or None,
                last_name=raw_iiko_snapshot.get("last_name") or None,
                position_name=iiko_position_name,
                position_code=iiko_position_code,
                staff_code=raw_iiko_snapshot.get("pin_code") or None,
                iiko_code=raw_iiko_snapshot.get("code") or None,
                iiko_id=raw_iiko_snapshot.get("id") or None,
                workplace_restaurant_id=getattr(primary_restaurant, "id", None),
                workplace_restaurant_name=getattr(primary_restaurant, "name", None),
                department_code=department_code,
                restaurant_ids=[int(row.id) for row in iiko_restaurants if row.id is not None],
                restaurant_names=[
                    row.name or f"Restaurant #{row.id}"
                    for row in iiko_restaurants
                    if row.id is not None
                ],
                department_codes=department_codes,
            )
    except IikoIntegrationError as exc:
        iiko_error = str(exc.detail)
    except Exception as exc:
        iiko_error = f"Unexpected iiko error: {exc}"
        logger.exception("Failed to load iiko sync preview for user %s", db_user.id)

    return EmployeeIikoSyncPreviewResponse(
        local=local_snapshot,
        iiko=iiko_snapshot,
        iiko_error=iiko_error,
    )


@router.post("/{user_id}/photo", response_model=PhotoUploadResponse, status_code=status.HTTP_201_CREATED)
async def upload_employee_photo(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    _ensure_can_edit_user(db, current_user, db_user)

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only image uploads are allowed")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    # Normalize to JPEG for broader browser support (e.g., HEIC/WebP uploads).
    upload_filename = file.filename or "photo.jpg"
    upload_content = content
    upload_content_type = file.content_type
    try:
        img = Image.open(BytesIO(content))
        img_format = (img.format or "").upper()
        if img_format != "JPEG" or upload_content_type not in {"image/jpeg", "image/jpg"}:
            if img.mode in {"RGBA", "LA"} or (img.mode == "P" and "transparency" in img.info):
                rgba = img.convert("RGBA")
                background = Image.new("RGB", rgba.size, (255, 255, 255))
                background.paste(rgba, mask=rgba.split()[-1])
                img = background
            else:
                img = img.convert("RGB")
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=90)
            upload_content = buf.getvalue()
            upload_content_type = "image/jpeg"
            base = os.path.splitext(upload_filename)[0] or "photo"
            upload_filename = f"{base}.jpg"
    except UnidentifiedImageError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported image format")
    except Exception:
        logger.warning("Failed to normalize employee photo; uploading original", exc_info=True)

    photo_key, photo_url = upload_employee_photo_with_url(
        user_id, upload_filename, upload_content, upload_content_type
    )
    db_user.photo_key = photo_key
    db.commit()
    db.refresh(db_user)
    return PhotoUploadResponse(photo_key=photo_key, photo_url=photo_url)


@router.put("/{user_id}", response_model=EmployeeCardPublic)
def update_employee(
    user_id: int,
    payload: EmployeeUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    db_user = (
        db.query(User)
        .options(
            joinedload(User.company),
            joinedload(User.role),
            joinedload(User.position).joinedload(Position.payment_format),
            joinedload(User.position).joinedload(Position.restaurant_subdivision),
            joinedload(User.restaurants),
        )
        .filter(User.id == user_id)
        .first()
    )
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    original_values = {
        "first_name": db_user.first_name,
        "last_name": db_user.last_name,
        "company": db_user.company,
        "position": db_user.position,
        "workplace": db_user.workplace_restaurant,
        "restaurants": list(db_user.restaurants or []),
        "rate": db_user.rate,
        "fired": db_user.fired,
    }

    fields_set = getattr(payload, "model_fields_set", set())
    sync_only_allowed_fields = {
        "first_name",
        "last_name",
        "middle_name",
        "staff_code",
        "iiko_code",
        "workplace_restaurant_id",
        "add_to_iiko",
        "iiko_sync_restaurant_id",
        "iiko_department_restaurant_ids",
    }
    can_edit_employee = _can_edit_user(db, current_user, db_user)
    if not can_edit_employee:
        if not payload.add_to_iiko:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: insufficient permissions to manage this employee",
            )
        _ensure_can_sync_user_iiko(db, current_user, db_user)
        disallowed_sync_fields = {field for field in fields_set if field not in sync_only_allowed_fields}
        if disallowed_sync_fields:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: insufficient permissions to edit employee fields",
            )

    rate_fields_requested = "rate" in fields_set or "individual_rate" in fields_set
    if rate_fields_requested and not has_permission(current_user, PermissionCode.STAFF_RATE_MANAGE):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: insufficient permissions to edit rate",
        )
    can_edit_rate = can_view_rate(current_user, db_user) and has_permission(
        current_user, PermissionCode.STAFF_RATE_MANAGE
    )
    rate_was_provided = "rate" in fields_set and can_edit_rate
    individual_rate_was_provided = "individual_rate" in fields_set and can_edit_rate

    if (
        "role_id" in fields_set
        and "position_id" not in fields_set
        and db_user.position is None
        and not has_permission(current_user, PermissionCode.STAFF_ROLES_ASSIGN)
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: insufficient permissions to edit role",
        )
    iiko_sync_restaurant_id = None
    if payload.iiko_sync_restaurant_id is not None:
        try:
            parsed_sync_restaurant_id = int(payload.iiko_sync_restaurant_id)
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid iiko_sync_restaurant_id")
        if parsed_sync_restaurant_id > 0:
            iiko_sync_restaurant_id = parsed_sync_restaurant_id

    iiko_department_restaurant_ids = None
    if payload.iiko_department_restaurant_ids is not None:
        iiko_department_restaurant_ids = []
        seen_department_restaurant_ids = set()
        for raw_id in payload.iiko_department_restaurant_ids:
            try:
                parsed_id = int(raw_id)
            except Exception:
                continue
            if parsed_id <= 0 or parsed_id in seen_department_restaurant_ids:
                continue
            seen_department_restaurant_ids.add(parsed_id)
            iiko_department_restaurant_ids.append(parsed_id)

    if payload.add_to_iiko:
        ensure_permissions(
            current_user,
            PermissionCode.STAFF_EMPLOYEES_IIKO_SYNC,
            PermissionCode.IIKO_MANAGE,
        )
        _validate_iiko_restaurant_scope(
            db,
            current_user,
            sync_restaurant_id=iiko_sync_restaurant_id,
            department_restaurant_ids=iiko_department_restaurant_ids,
        )

    if payload.first_name is not None:
        db_user.first_name = payload.first_name
    if payload.last_name is not None:
        db_user.last_name = payload.last_name
    if payload.middle_name is not None:
        db_user.middle_name = payload.middle_name
    if payload.gender is not None:
        db_user.gender = payload.gender
    if payload.staff_code is not None:
        db_user.staff_code = payload.staff_code
    if payload.iiko_code is not None:
        db_user.iiko_code = payload.iiko_code
    if payload.iiko_id is not None:
        if not _can_edit_iiko_id(current_user):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: insufficient permissions to edit iiko_id",
            )
        db_user.iiko_id = (payload.iiko_id or "").strip() or None
    if payload.phone_number is not None:
        db_user.phone_number = (payload.phone_number or "").strip() or None
    if payload.email is not None:
        db_user.email = (payload.email or "").strip() or None
    if rate_was_provided:
        db_user.rate = payload.rate
    if "hire_date" in fields_set:
        db_user.hire_date = payload.hire_date
    if "fire_date" in fields_set:
        db_user.fire_date = payload.fire_date
    if payload.fired is not None:
        db_user.fired = payload.fired
    if "birth_date" in fields_set:
        db_user.birth_date = payload.birth_date
    if payload.photo_key is not None:
        db_user.photo_key = payload.photo_key
    if payload.is_cis_employee is not None:
        db_user.is_cis_employee = payload.is_cis_employee
    if payload.is_formalized is not None:
        db_user.is_formalized = payload.is_formalized
    if payload.confidential_data_consent is not None:
        db_user.confidential_data_consent = bool(payload.confidential_data_consent)
    if payload.password:
        db_user.hashed_password = hash_password(payload.password)

    if payload.company_id is not None:
        if payload.company_id:
            company = db.query(Company).filter(Company.id == payload.company_id).first()
            if not company:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Company not found")
            db_user.company = company
        else:
            db_user.company = None

    if payload.position_id is not None:
        if payload.position_id:
            position = db.query(Position).filter(Position.id == payload.position_id).first()
            if not position:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Position not found")
            db_user.position = position
            if not rate_was_provided:
                db_user.rate = position.rate
        else:
            db_user.position = None
            if not rate_was_provided:
                db_user.rate = None
            db_user.role = None

    if payload.role_id is not None and "position_id" not in fields_set and db_user.position is None:
        if payload.role_id:
            role = db.query(Role).filter(Role.id == payload.role_id).first()
            if not role:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Role not found")
            db_user.role = role
        else:
            db_user.role = None

    if db_user.position is not None:
        db_user.role = db_user.position.role

    if individual_rate_was_provided:
        if payload.individual_rate is None:
            db_user.individual_rate = None
            if not rate_was_provided:
                position = db_user.position
                if position and position.rate is not None:
                    db_user.rate = position.rate
                else:
                    db_user.rate = None
        else:
            db_user.individual_rate = payload.individual_rate
            db_user.rate = payload.individual_rate

    if payload.workplace_restaurant_id is not None:
        if payload.workplace_restaurant_id:
            workplace_restaurant = (
                db.query(Restaurant)
                .filter(Restaurant.id == payload.workplace_restaurant_id)
                .first()
            )
            if not workplace_restaurant:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Restaurant not found")
            db_user.workplace_restaurant = workplace_restaurant
        else:
            db_user.workplace_restaurant = None

    if payload.restaurant_ids is not None:
        if payload.restaurant_ids:
            restaurants = db.query(Restaurant).filter(Restaurant.id.in_(payload.restaurant_ids)).all()
            db_user.restaurants = restaurants
        else:
            if getattr(payload, "clear_restaurants", False):
                db_user.restaurants = []
            else:
                if db_user.restaurants:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="restaurant_ids is empty; set clear_restaurants=true to remove access",
                    )

    if "company_id" not in fields_set:
        inferred_company_id = None
        workplace_company_id = getattr(db_user.workplace_restaurant, "company_id", None)
        if workplace_company_id:
            inferred_company_id = workplace_company_id
        elif db_user.restaurants:
            company_ids = {
                getattr(item, "company_id", None)
                for item in (db_user.restaurants or [])
                if getattr(item, "company_id", None)
            }
            if len(company_ids) == 1:
                inferred_company_id = next(iter(company_ids))
        if inferred_company_id:
            company = db.query(Company).filter(Company.id == inferred_company_id).first()
            if company:
                db_user.company = company

    new_row_id = build_employee_row_id(
        last_name=db_user.last_name,
        first_name=db_user.first_name,
        middle_name=db_user.middle_name,
        birth_date=db_user.birth_date,
    )
    if new_row_id != getattr(db_user, "employee_row_id", None):
        if new_row_id:
            conflict_id = (
                db.query(User.id)
                .filter(User.employee_row_id == new_row_id, User.id != db_user.id)
                .scalar()
            )
            if conflict_id:
                conflict_user = (
                    db.query(User)
                    .options(
                        joinedload(User.position),
                        joinedload(User.workplace_restaurant),
                    )
                    .filter(User.id == conflict_id)
                    .first()
                )
                detail = {"code": "employee_duplicate"}
                if conflict_user:
                    can_open_card = _can_view_duplicate_employee(db, current_user, conflict_user)
                    detail["can_open_card"] = can_open_card
                    detail["employee"] = _build_duplicate_employee_payload(
                        db,
                        conflict_user,
                        include_id=can_open_card,
                    )
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detail)
        db_user.employee_row_id = new_row_id

    changes = []
    if original_values["first_name"] != db_user.first_name:
        changes.append(
            {
                "field": "first_name",
                "old_value": original_values["first_name"],
                "new_value": db_user.first_name,
            }
        )
    if original_values["last_name"] != db_user.last_name:
        changes.append(
            {
                "field": "last_name",
                "old_value": original_values["last_name"],
                "new_value": db_user.last_name,
            }
        )
    if getattr(original_values["company"], "id", None) != getattr(db_user.company, "id", None):
        changes.append(
            {
                "field": "company",
                "old_value": format_ref(original_values["company"]),
                "new_value": format_ref(db_user.company),
            }
        )
    if getattr(original_values["position"], "id", None) != getattr(db_user.position, "id", None):
        changes.append(
            {
                "field": "position",
                "old_value": format_ref(original_values["position"]),
                "new_value": format_ref(db_user.position),
            }
        )
    if getattr(original_values["workplace"], "id", None) != getattr(db_user.workplace_restaurant, "id", None):
        changes.append(
            {
                "field": "workplace_restaurant",
                "old_value": format_ref(original_values["workplace"]),
                "new_value": format_ref(db_user.workplace_restaurant),
            }
        )
    if _normalize_rate_value(original_values["rate"]) != _normalize_rate_value(db_user.rate):
        changes.append(
            {
                "field": "rate",
                "old_value": original_values["rate"],
                "new_value": db_user.rate,
            }
        )
    original_restaurant_ids = {
        item.id for item in (original_values["restaurants"] or []) if item and item.id is not None
    }
    updated_restaurant_ids = {
        item.id for item in (db_user.restaurants or []) if item and item.id is not None
    }
    if original_restaurant_ids != updated_restaurant_ids:
        changes.append(
            {
                "field": "restaurants",
                "old_value": _format_restaurant_refs(original_values["restaurants"]),
                "new_value": _format_restaurant_refs(db_user.restaurants),
            }
        )
    if original_values["fired"] != db_user.fired:
        changes.append(
            {
                "field": "fired",
                "old_value": original_values["fired"],
                "new_value": db_user.fired,
            }
        )

    log_employee_changes(
        db,
        user_id=db_user.id,
        changed_by_id=current_user.id,
        changes=changes,
    )

    db.commit()
    db.refresh(db_user)
    if payload.add_to_iiko:
        try:
            previous_iiko_code = (getattr(db_user, "iiko_code", None) or "").strip() or None
            previous_iiko_id = (getattr(db_user, "iiko_id", None) or "").strip() or None
            iiko_uid = add_user_to_iiko(
                db,
                db_user,
                sync_restaurant_id=iiko_sync_restaurant_id,
                department_restaurant_ids=iiko_department_restaurant_ids,
            )
            if iiko_uid:
                db_user.iiko_id = iiko_uid
            updated_iiko_code = (getattr(db_user, "iiko_code", None) or "").strip() or None
            updated_iiko_id = (getattr(db_user, "iiko_id", None) or "").strip() or None
            if updated_iiko_code != previous_iiko_code or updated_iiko_id != previous_iiko_id:
                db.commit()
                db.refresh(db_user)
            setattr(db_user, "iiko_sync_error", None)
        except IikoIntegrationError as exc:
            setattr(db_user, "iiko_sync_error", exc.detail)
            logger.warning("iiko sync failed for user %s: %s", db_user.id, exc.detail)
        except Exception as exc:
            setattr(db_user, "iiko_sync_error", f"Unexpected iiko error: {exc}")
            logger.exception("Failed to add user %s to iiko", db_user.id)
    return _to_employee_card(db_user, current_user)


@router.delete("/{user_id}")
async def delete_employee(
    user_id: int,
    request: Request,
    delete_from_iiko: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    _ensure_can_edit_user(db, current_user, db_user)
    iiko_error = None
    if delete_from_iiko:
        try:
            fire_user_in_iiko(db, db_user)
        except IikoIntegrationError as exc:
            iiko_error = exc.detail
            logger.warning("iiko delete failed for user %s: %s", db_user.id, exc.detail)
        except Exception as exc:
            iiko_error = f"Unexpected iiko error: {exc}"
            logger.exception("Failed to delete user %s from iiko", db_user.id)
    try:
        payload = await request.json()
        comment = (payload or {}).get("comment") or ""
    except Exception:
        comment = ""
    comment = comment.strip()
    if not comment:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Комментарий обязателен")

    if not db_user.fired:
        old_value = db_user.fired
        db_user.fired = True
        log_employee_changes(
            db,
            user_id=db_user.id,
            changed_by_id=current_user.id,
            changes=[
                {
                    "field": "fired",
                    "old_value": old_value,
                    "new_value": True,
                    "comment": comment,
                    "source": "delete_employee",
                }
            ],
        )
    db_user.fire_date = date.today()
    db.commit()
    return {"ok": True, "iiko_sync_error": iiko_error}


@router.post("/{user_id}/restore", response_model=EmployeeCardPublic)
async def restore_employee(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> EmployeeCardPublic:
    ensure_permissions(current_user, PermissionCode.STAFF_EMPLOYEES_RESTORE)
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    _ensure_can_edit_user(db, current_user, db_user)

    try:
        payload = await request.json()
        comment = (payload or {}).get("comment") or ""
    except Exception:
        comment = ""
    comment = comment.strip()

    if db_user.fired:
        old_value = db_user.fired
        db_user.fired = False
        log_employee_changes(
            db,
            user_id=db_user.id,
            changed_by_id=current_user.id,
            changes=[
                {
                    "field": "fired",
                    "old_value": old_value,
                    "new_value": False,
                    "comment": comment,
                    "source": "restore_employee",
                }
            ],
        )
    if db_user.fire_date is not None:
        db_user.fire_date = None
    db.commit()
    db.refresh(db_user)
    return _to_employee_card(db_user, current_user)
